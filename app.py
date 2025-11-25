from __future__ import annotations

import json
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from flask import Flask, abort, jsonify, redirect, render_template, request, url_for
from openpyxl import Workbook, load_workbook

PATIENT_FILE = Path("patients.xlsx")
SHEET_NAME = "Patients"
DOCTOR_FILE = Path("doctors.xlsx")
DOCTOR_SHEET_NAME = "Doctors"
OPD_FILE = Path("opd.xlsx")
OPD_SHEET_NAME = "OPD"
ADMISSION_FILE = Path("admissions.xlsx")
ADMISSION_SHEET_NAME = "Admissions"
CHARGE_FILE = Path("charges.xlsx")
CHARGE_SHEET_NAME = "Charges"
BILLING_FILE = Path("billing.xlsx")
BILLING_SHEET_NAME = "Billing"
ADMISSION_CHARGES_FILE = Path("admission_charges.xlsx")
ADMISSION_CHARGES_SHEET_NAME = "ChargeEntries"

FIELD_SECTIONS = [
    {
        "key": "personal",
        "label": "1. Personal Information",
        "fields": [
            {"name": "full_name", "label": "Full Name", "type": "text", "required": True},
            {"name": "gender", "label": "Gender", "type": "select", "options": ["Male", "Female", "Other"], "required": True},
            {"name": "dob", "label": "Date of Birth", "type": "date"},
            {"name": "age", "label": "Age", "type": "number", "min": 0},
            {"name": "blood_group", "label": "Blood Group", "type": "text"},
            {"name": "marital_status", "label": "Marital Status", "type": "select", "options": ["Single", "Married", "Divorced", "Widowed", "Other"]},
            {"name": "photo_url", "label": "Photo URL", "type": "text"},
        ],
    },
    {
        "key": "contact",
        "label": "2. Contact Details",
        "fields": [
            {"name": "mobile_primary", "label": "Mobile Number (Primary)", "type": "tel", "required": True},
            {"name": "mobile_alternate", "label": "Alternate Number", "type": "tel"},
            {"name": "email", "label": "Email", "type": "email"},
            {"name": "address_permanent", "label": "Permanent Address", "type": "textarea"},
            {"name": "address_local", "label": "Local Address", "type": "textarea"},
        ],
    },
    {
        "key": "identification",
        "label": "3. Identification Details",
        "fields": [
            {"name": "aadhar_number", "label": "Aadhar Number", "type": "text"},
            {"name": "pan_number", "label": "PAN", "type": "text"},
            {"name": "hospital_id", "label": "Patient Unique Hospital ID (MRN)", "type": "text", "auto": True},
            {"name": "govt_id_upload", "label": "Government ID Upload (URL / reference)", "type": "text"},
        ],
    },
    {
        "key": "emergency",
        "label": "4. Emergency Contact Details",
        "fields": [
            {"name": "emergency_name", "label": "Name of Emergency Contact", "type": "text"},
            {"name": "emergency_relationship", "label": "Relationship", "type": "text"},
            {"name": "emergency_mobile", "label": "Mobile Number", "type": "tel"},
            {"name": "emergency_address", "label": "Address", "type": "textarea"},
        ],
    },
    {
        "key": "medical",
        "label": "6. Medical Background",
        "fields": [
            {"name": "allergies", "label": "Allergies", "type": "textarea"},
            {"name": "existing_conditions", "label": "Existing Conditions", "type": "textarea"},
            {"name": "past_surgeries", "label": "Past Surgeries / Hospitalizations", "type": "textarea"},
            {"name": "current_medication", "label": "Current Medication", "type": "textarea"},
            {"name": "family_history", "label": "Family Medical History", "type": "textarea"},
            {"name": "habits", "label": "Smoking/Alcohol Habits", "type": "textarea"},
        ],
    },
    {
        "key": "insurance",
        "label": "7. Insurance / Billing Info",
        "fields": [
            {"name": "billing_type", "label": "Billing Type", "type": "select", "options": ["Cash", "Insurance", "Third-party / Corporate"], "required": True},
            {"name": "insurance_provider", "label": "Insurance Provider", "type": "text"},
            {"name": "policy_number", "label": "Policy Number", "type": "text"},
            {"name": "tpa_details", "label": "TPA Details", "type": "textarea"},
            {"name": "claim_type", "label": "Claim Type", "type": "select", "options": ["Cashless", "Reimbursement"]},
        ],
    },
    {
        "key": "consent",
        "label": "8. Consent & Legal",
        "fields": [
            {"name": "treatment_consent", "label": "Treatment Consent", "type": "checkbox"},
            {"name": "surgery_consent", "label": "Surgery Consent", "type": "checkbox"},
            {"name": "privacy_consent", "label": "Privacy / PDPA consent", "type": "checkbox"},
            {"name": "guardian_details", "label": "Guardian/Parent Details (for minors)", "type": "textarea"},
        ],
    },
]

DOCTOR_FIELD_SECTIONS = [
    {
        "key": "personal",
        "label": "1. Personal Details",
        "fields": [
            {"name": "full_name", "label": "Full Name", "type": "text", "required": True},
            {"name": "gender", "label": "Gender", "type": "select", "options": ["Male", "Female", "Other"], "required": True},
            {"name": "dob", "label": "Date of Birth", "type": "date"},
            {"name": "age", "label": "Age", "type": "number", "min": 0},
            {"name": "contact_primary", "label": "Contact Number (Primary)", "type": "tel", "required": True},
            {"name": "contact_alternate", "label": "Alternate Number", "type": "tel"},
            {"name": "email", "label": "Email", "type": "email"},
            {"name": "address", "label": "Address", "type": "textarea"},
            {"name": "photo_url", "label": "Photo URL", "type": "text"},
        ],
    },
    {
        "key": "professional",
        "label": "2. Professional / Qualification Details",
        "fields": [
            {"name": "qualification", "label": "Qualification", "type": "select", "options": ["MBBS", "MD", "MS", "DNB", "BDS", "Other"], "required": True},
            {"name": "specialization", "label": "Specialization", "type": "text", "required": True},
            {"name": "sub_specialization", "label": "Sub-specialization", "type": "text"},
            {"name": "registration_number", "label": "Registration Number (Medical Council)", "type": "text", "required": True},
            {"name": "experience_years", "label": "Experience (Years)", "type": "number", "min": 0},
            {"name": "fellowships", "label": "Fellowships / Additional Certifications", "type": "textarea"},
        ],
    },
    {
        "key": "employment",
        "label": "3. Employment / Hospital Details",
        "fields": [
            {"name": "doctor_type", "label": "Doctor Type", "type": "select", "options": ["Visiting", "Full-time", "Consultant", "Surgeon", "Duty Doctor"], "required": True},
            {"name": "department", "label": "Department", "type": "select", "options": ["General Medicine", "Surgery", "Orthopedics", "Neurology", "Pediatrics", "Cardiology", "Gynecology", "Dermatology", "ENT", "Ophthalmology", "Other"], "required": True},
            {"name": "joining_date", "label": "Joining Date", "type": "date"},
            {"name": "status", "label": "Status", "type": "select", "options": ["Active", "Inactive"], "required": True},
        ],
    },
    {
        "key": "consultation",
        "label": "4. Consultation & Service Details",
        "fields": [
            {"name": "opd_fee_initial", "label": "OPD Consultation Fee (Initial)", "type": "number", "min": 0},
            {"name": "opd_fee_followup", "label": "Follow-up Fee", "type": "number", "min": 0},
            {"name": "opd_days", "label": "OPD Days (e.g., Mon,Wed,Fri)", "type": "text"},
            {"name": "opd_timings", "label": "OPD Timings (e.g., 9AM-12PM, 5PM-8PM)", "type": "text"},
            {"name": "opd_max_patients", "label": "Max Patients per day", "type": "number", "min": 0},
            {"name": "ipd_visit_charge", "label": "IPD Visit Charge", "type": "number", "min": 0},
            {"name": "icu_visit_charge", "label": "ICU Visit Charge", "type": "number", "min": 0},
            {"name": "round_frequency", "label": "Round Visit Frequency", "type": "select", "options": ["Once Daily", "Twice Daily", "As Required"]},
            {"name": "surgery_fee", "label": "Surgery Fee", "type": "number", "min": 0},
            {"name": "ot_eligibility", "label": "OT Eligibility", "type": "checkbox"},
            {"name": "surgery_types", "label": "Surgery Types Handled", "type": "textarea"},
        ],
    },
    {
        "key": "availability",
        "label": "5. Availability / Scheduling",
        "fields": [
            {"name": "weekly_schedule", "label": "Weekly Availability Schedule", "type": "textarea"},
            {"name": "locations", "label": "Locations (multiple branches)", "type": "textarea"},
            {"name": "emergency_oncall", "label": "Emergency On-Call Availability", "type": "checkbox"},
            {"name": "leaves_offdays", "label": "Leaves / Off days", "type": "textarea"},
        ],
    },
    {
        "key": "banking",
        "label": "6. Banking & Payroll",
        "fields": [
            {"name": "bank_account", "label": "Bank Account Number", "type": "text"},
            {"name": "ifsc_code", "label": "IFSC Code", "type": "text"},
            {"name": "pan_number", "label": "PAN Number", "type": "text"},
            {"name": "upi_id", "label": "UPI ID", "type": "text"},
            {"name": "payment_mode", "label": "Payment Mode", "type": "select", "options": ["Cash", "Bank Transfer", "UPI"]},
        ],
    },
    {
        "key": "login",
        "label": "7. Login / Access",
        "fields": [
            {"name": "username", "label": "Username", "type": "text"},
            {"name": "password", "label": "Password (hashed)", "type": "password"},
            {"name": "role", "label": "Role", "type": "select", "options": ["Doctor"], "required": False},
            {"name": "access_level", "label": "Access Level", "type": "select", "options": ["OPD-only", "IPD-only", "Full access"]},
        ],
    },
]

OPD_FIELD_SECTIONS = [
    {
        "key": "basic",
        "label": "1. Basic Patient Information",
        "fields": [
            {"name": "patient_name", "label": "Patient Name", "type": "text", "required": True},
            {"name": "age", "label": "Age", "type": "number", "min": 0},
            {"name": "gender", "label": "Gender", "type": "select", "options": ["Male", "Female", "Other"], "required": True},
            {"name": "mobile_number", "label": "Mobile Number", "type": "tel", "required": True},
            {"name": "opd_id", "label": "OPD ID / Token No", "type": "text", "auto": True},
        ],
    },
    {
        "key": "visit",
        "label": "2. Visit Details",
        "fields": [
            {"name": "visit_date_time", "label": "Visit Date & Time", "type": "datetime-local", "required": True},
            {"name": "department", "label": "Department", "type": "select", "options": ["General Medicine", "Surgery", "Orthopedics", "Neurology", "Pediatrics", "Cardiology", "Gynecology", "Dermatology", "ENT", "Ophthalmology", "Other"], "required": True},
            {"name": "doctor_name", "label": "Doctor Name / Doctor ID", "type": "doctor_search", "required": True},
            {"name": "visit_type", "label": "Visit Type", "type": "select", "options": ["New Visit", "Follow-up Visit"], "required": True},
        ],
    },
    {
        "key": "clinical",
        "label": "3. Clinical Details",
        "fields": [
            {"name": "chief_complaint", "label": "Chief Complaint", "type": "textarea", "required": True},
            {"name": "diagnosis", "label": "Diagnosis", "type": "textarea"},
            {"name": "treatment_for", "label": "Treatment For", "type": "text"},
        ],
    },
    {
        "key": "billing",
        "label": "4. Billing Details",
        "fields": [
            {"name": "consultation_fee", "label": "Consultation Fee", "type": "number", "min": 0, "required": True},
            {"name": "discount", "label": "Discount", "type": "number", "min": 0},
            {"name": "payment_method", "label": "Payment Method", "type": "select", "options": ["Cash", "UPI", "Card"], "required": True},
            {"name": "bill_number", "label": "Bill Number", "type": "text", "auto": True},
        ],
    },
    {
        "key": "optional",
        "label": "5. Optional Information",
        "fields": [
            {"name": "address", "label": "Address", "type": "textarea"},
            {"name": "allergies", "label": "Allergies", "type": "textarea"},
            {"name": "existing_conditions", "label": "Existing Conditions (Diabetes/BP)", "type": "textarea"},
            {"name": "insurance", "label": "Insurance?", "type": "select", "options": ["Yes", "No"]},
        ],
    },
]

ADMISSION_FIELD_SECTIONS = [
    {
        "key": "header",
        "label": "1. Admission Header",
        "fields": [
            {"name": "admission_id", "label": "Admission ID", "type": "text", "auto": True},
            {"name": "admission_date_time", "label": "Admission Date & Time", "type": "datetime-local", "required": True},
            {"name": "admission_type", "label": "Admission Type", "type": "select", "options": ["Planned / Elective", "Emergency", "Transfer from another hospital"], "required": True},
            {"name": "visit_type", "label": "Visit Type", "type": "select", "options": ["Inpatient (IPD)", "Day Care"], "required": True},
        ],
    },
    {
        "key": "patient",
        "label": "2. Patient & Attendant Details",
        "fields": [
            {"name": "patient_id", "label": "Patient ID (from Patient master)", "type": "patient_search", "required": True},
            {"name": "patient_name", "label": "Patient Name", "type": "text", "required": True},
            {"name": "patient_age", "label": "Age", "type": "text"},
            {"name": "patient_gender", "label": "Gender", "type": "text"},
            {"name": "attendant_name", "label": "Attendant / Guardian Name", "type": "text", "required": True},
            {"name": "attendant_relationship", "label": "Relationship to Patient", "type": "select", "options": ["Father", "Mother", "Spouse", "Son", "Daughter", "Other"], "required": True},
            {"name": "attendant_mobile", "label": "Attendant Mobile Number", "type": "tel", "required": True},
            {"name": "attendant_address", "label": "Attendant Address", "type": "textarea"},
        ],
    },
    {
        "key": "clinical",
        "label": "3. Clinical Intake / Reason for Admission",
        "fields": [
            {"name": "chief_complaint", "label": "Chief Complaint", "type": "textarea", "required": True},
            {"name": "provisional_diagnosis", "label": "Provisional Diagnosis", "type": "textarea"},
            {"name": "mode_of_arrival", "label": "Mode of Arrival", "type": "select", "options": ["Walk-in", "Wheelchair", "Stretcher", "Ambulance"], "required": True},
            {"name": "referring_doctor", "label": "Referring Doctor/Hospital", "type": "text"},
            {"name": "triage_category", "label": "Triage Category (for emergency)", "type": "select", "options": ["Red", "Yellow", "Green"]},
        ],
    },
    {
        "key": "doctor",
        "label": "4. Doctor & Department Mapping",
        "fields": [
            {"name": "admitting_department", "label": "Admitting Department", "type": "select", "options": ["General Medicine", "Surgery", "Orthopedics", "Neurology", "Pediatrics", "Cardiology", "Gynecology", "Dermatology", "ENT", "Ophthalmology", "ICU", "NICU", "PICU", "Other"], "required": True},
            {"name": "admitting_consultant", "label": "Admitting Consultant Doctor ID", "type": "doctor_search", "required": True},
            {"name": "assistant_doctor", "label": "Assistant Doctor / Duty Doctor", "type": "doctor_search"},
            {"name": "clinical_unit", "label": "Clinical Unit / Team", "type": "text"},
        ],
    },
    {
        "key": "bed",
        "label": "5. Bed / Room / Ward Allocation",
        "fields": [
            {"name": "ward", "label": "Ward", "type": "select", "options": ["General", "Private", "Semi-Private", "ICU", "NICU", "PICU", "Isolation", "Other"], "required": True},
            {"name": "room_type", "label": "Room Type", "type": "select", "options": ["General", "Twin Sharing", "Deluxe", "Suite", "ICU Bed", "Other"], "required": True},
            {"name": "room_number", "label": "Room Number", "type": "text", "required": True},
            {"name": "bed_number", "label": "Bed Number", "type": "text", "required": True},
            {"name": "bed_status", "label": "Bed Status", "type": "select", "options": ["Occupied", "Reserved", "Blocked"], "required": True},
            {"name": "expected_stay", "label": "Expected Length of Stay (days)", "type": "number", "min": 0},
        ],
    },
    {
        "key": "billing",
        "label": "6. Billing & Payment Info",
        "fields": [
            {"name": "billing_category", "label": "Billing Category / Payer Type", "type": "select", "options": ["Self / Cash", "Insurance", "Corporate / PSU / TPA", "Government Scheme"], "required": True},
            {"name": "tariff_plan", "label": "Tariff Plan / Package", "type": "text"},
            {"name": "initial_deposit", "label": "Initial Deposit Amount", "type": "number", "min": 0},
            {"name": "payment_mode", "label": "Mode of Payment", "type": "select", "options": ["Cash", "Card", "UPI", "NEFT"]},
            {"name": "receipt_number", "label": "Receipt / Transaction Number", "type": "text"},
        ],
    },
    {
        "key": "insurance",
        "label": "7. Insurance / TPA Section",
        "fields": [
            {"name": "insurance_company", "label": "Insurance Company Name", "type": "text"},
            {"name": "tpa_name", "label": "TPA Name", "type": "text"},
            {"name": "policy_number", "label": "Policy Number", "type": "text"},
            {"name": "policy_holder_name", "label": "Policy Holder Name", "type": "text"},
            {"name": "policy_holder_relation", "label": "Policy Holder Relation", "type": "text"},
            {"name": "policy_validity_from", "label": "Policy Validity From", "type": "date"},
            {"name": "policy_validity_to", "label": "Policy Validity To", "type": "date"},
            {"name": "authorization_number", "label": "Authorization / Pre-Auth Number", "type": "text"},
            {"name": "approved_amount", "label": "Approved Amount", "type": "number", "min": 0},
            {"name": "corporate_name", "label": "Corporate Name", "type": "text"},
        ],
    },
    {
        "key": "consent",
        "label": "8. Consent & Legal",
        "fields": [
            {"name": "treatment_consent", "label": "General Treatment Consent", "type": "checkbox"},
            {"name": "icu_consent", "label": "ICU Consent", "type": "checkbox"},
            {"name": "surgery_consent", "label": "Surgery Consent", "type": "checkbox"},
            {"name": "dnr_flag", "label": "DNR / Advance Directive", "type": "checkbox"},
            {"name": "consent_taken_from", "label": "Consent Taken From", "type": "select", "options": ["Patient", "Relative"]},
            {"name": "consent_relative_name", "label": "Relative Name & Relation", "type": "text"},
            {"name": "consent_form_upload", "label": "Consent Form Upload (file reference)", "type": "text"},
        ],
    },
    {
        "key": "operational",
        "label": "9. Operational / Admin Details",
        "fields": [
            {"name": "created_by", "label": "Admission Created By (User/Employee ID)", "type": "text"},
            {"name": "created_date_time", "label": "Created Date & Time", "type": "datetime-local"},
            {"name": "last_updated_by", "label": "Last Updated By", "type": "text"},
            {"name": "status", "label": "Status", "type": "select", "options": ["Admitted", "Cancelled", "Discharged", "Transferred"], "required": True},
            {"name": "source_of_admission", "label": "Source of Admission", "type": "select", "options": ["OPD → IPD", "Emergency → IPD", "Direct Admission", "Transfer from another hospital"], "required": True},
        ],
    },
    {
        "key": "tracking",
        "label": "10. During-Stay Tracking",
        "fields": [
            {"name": "current_ward", "label": "Current Ward", "type": "text"},
            {"name": "current_bed", "label": "Current Bed", "type": "text"},
            {"name": "transfer_history", "label": "Transfer History", "type": "textarea"},
            {"name": "clinical_severity", "label": "Clinical Severity / Risk Category", "type": "text"},
            {"name": "isolation_flag", "label": "Isolation / Infection Flag", "type": "select", "options": ["Yes", "No"]},
        ],
    },
    {
        "key": "discharge",
        "label": "11. Discharge-Related Fields",
        "fields": [
            {"name": "discharge_date_time", "label": "Discharge Date & Time", "type": "datetime-local"},
            {"name": "discharge_type", "label": "Discharge Type", "type": "select", "options": ["Discharged to home", "LAMA / DAMA", "Referred to other hospital", "Death"]},
            {"name": "final_diagnosis", "label": "Final Diagnosis", "type": "textarea"},
            {"name": "condition_at_discharge", "label": "Condition at Discharge", "type": "select", "options": ["Stable", "Critical", "Improved"]},
            {"name": "followup_date", "label": "Follow-up Date", "type": "date"},
            {"name": "followup_doctor", "label": "Follow-up Doctor", "type": "text"},
            {"name": "followup_department", "label": "Follow-up Department", "type": "text"},
        ],
    },
]

CHARGE_FIELD_SECTIONS = [
    {
        "key": "registration",
        "label": "1. Registration & Administrative Charges",
        "fields": [
            {"name": "registration_fee", "label": "Registration Fee (₹)", "type": "number", "min": 0},
            {"name": "file_opening_charges", "label": "File Opening Charges (₹)", "type": "number", "min": 0},
            {"name": "card_opd_slip_charges", "label": "Card / OPD Slip Charges (₹)", "type": "number", "min": 0},
            {"name": "admission_processing_fee", "label": "Admission Processing Fee (₹)", "type": "number", "min": 0},
            {"name": "emergency_registration_fee", "label": "Emergency Registration Fee (₹)", "type": "number", "min": 0},
        ],
    },
    {
        "key": "room_bed",
        "label": "2. Room / Bed Charges (IPD) - Per Day",
        "fields": [
            {"name": "general_ward_bed", "label": "General Ward Bed Charges (₹)", "type": "number", "min": 0},
            {"name": "semi_private_room", "label": "Semi-Private Room (₹)", "type": "number", "min": 0},
            {"name": "private_room", "label": "Private Room (₹)", "type": "number", "min": 0},
            {"name": "deluxe_room", "label": "Deluxe Room (₹)", "type": "number", "min": 0},
            {"name": "suite_room", "label": "Suite Room (₹)", "type": "number", "min": 0},
            {"name": "icu", "label": "ICU (₹)", "type": "number", "min": 0},
            {"name": "iccu", "label": "ICCU (₹)", "type": "number", "min": 0},
            {"name": "nicu_picu", "label": "NICU / PICU (₹)", "type": "number", "min": 0},
            {"name": "ventilator_bed", "label": "Ventilator Bed Charges (₹)", "type": "number", "min": 0},
            {"name": "isolation_room", "label": "Isolation Room Charges (₹)", "type": "number", "min": 0},
        ],
    },
    {
        "key": "nursing",
        "label": "3. Nursing Charges",
        "fields": [
            {"name": "nursing_care_charge", "label": "Nursing Care Charge (₹)", "type": "number", "min": 0},
            {"name": "special_nursing_charge", "label": "Special Nursing Charge (₹)", "type": "number", "min": 0},
            {"name": "attendant_charges", "label": "Attendant Charges (₹)", "type": "number", "min": 0},
        ],
    },
    {
        "key": "doctor_visit",
        "label": "4. Doctor Visit Charges",
        "fields": [
            {"name": "opd_consultation_fee", "label": "OPD Consultation Fee (₹)", "type": "number", "min": 0},
            {"name": "opd_followup_fee", "label": "OPD Follow-up Fee (₹)", "type": "number", "min": 0},
            {"name": "ipd_daily_visit_charge", "label": "IPD Daily Visit Charge (₹)", "type": "number", "min": 0},
            {"name": "icu_visit_charge", "label": "ICU Visit Charge (₹)", "type": "number", "min": 0},
            {"name": "night_visit_charge", "label": "Night Visit Charge (₹)", "type": "number", "min": 0},
            {"name": "surgeon_visit_charge", "label": "Surgeon Visit Charge (₹)", "type": "number", "min": 0},
        ],
    },
    {
        "key": "procedures",
        "label": "5. Procedures & Treatment Charges",
        "fields": [
            {"name": "dressing", "label": "Dressing (₹)", "type": "number", "min": 0},
            {"name": "nebulization", "label": "Nebulization (₹)", "type": "number", "min": 0},
            {"name": "catheterization", "label": "Catheterization (₹)", "type": "number", "min": 0},
            {"name": "injection_charges", "label": "Injection Charges (IV/IM/SC) (₹)", "type": "number", "min": 0},
            {"name": "iv_fluids_administration", "label": "IV Fluids Administration (₹)", "type": "number", "min": 0},
            {"name": "enema", "label": "Enema (₹)", "type": "number", "min": 0},
            {"name": "blood_transfusion", "label": "Blood Transfusion Service Charge (₹)", "type": "number", "min": 0},
            {"name": "plaster_pop", "label": "Plaster / POP (₹)", "type": "number", "min": 0},
            {"name": "wound_suturing", "label": "Wound Suturing (₹)", "type": "number", "min": 0},
            {"name": "physiotherapy_session", "label": "Physiotherapy Session (₹)", "type": "number", "min": 0},
            {"name": "dialysis_session", "label": "Dialysis Session (₹)", "type": "number", "min": 0},
        ],
    },
    {
        "key": "ot_procedure",
        "label": "6. OT / Procedure Room Charges",
        "fields": [
            {"name": "ot_charges", "label": "OT Charges (₹)", "type": "number", "min": 0},
            {"name": "minor_ot_charges", "label": "Minor OT Charges (₹)", "type": "number", "min": 0},
            {"name": "anesthesia_charges", "label": "Anesthesia Charges (₹)", "type": "number", "min": 0},
            {"name": "anesthetist_visit_charge", "label": "Anesthetist Visit Charge (₹)", "type": "number", "min": 0},
            {"name": "surgeon_fee", "label": "Surgeon Fee (₹)", "type": "number", "min": 0},
            {"name": "assistant_surgeon_fee", "label": "Assistant Surgeon Fee (₹)", "type": "number", "min": 0},
            {"name": "recovery_room_charges", "label": "Recovery Room Charges (₹)", "type": "number", "min": 0},
        ],
    },
    {
        "key": "pharmacy",
        "label": "7. Pharmacy / Medicine Charges",
        "fields": [
            {"name": "tablets_charge", "label": "Tablets (₹)", "type": "number", "min": 0},
            {"name": "injections_charge", "label": "Injections (₹)", "type": "number", "min": 0},
            {"name": "iv_fluids_charge", "label": "IV Fluids (₹)", "type": "number", "min": 0},
            {"name": "consumables_charge", "label": "Consumables (cotton, syringe, cannula, gloves) (₹)", "type": "number", "min": 0},
            {"name": "surgical_consumables_charge", "label": "Surgical Consumables (sutures, drapes, implants) (₹)", "type": "number", "min": 0},
        ],
    },
    {
        "key": "bedside",
        "label": "8. Bedside Services",
        "fields": [
            {"name": "oxygen_charges", "label": "Oxygen Charges (per hour/cylinder) (₹)", "type": "number", "min": 0},
            {"name": "ventilator_charges", "label": "Ventilator Charges (₹)", "type": "number", "min": 0},
            {"name": "defibrillator_usage", "label": "Defibrillator Usage (₹)", "type": "number", "min": 0},
            {"name": "cpap_bipap_use", "label": "CPAP/BIPAP Use (₹)", "type": "number", "min": 0},
            {"name": "suction_machine", "label": "Suction Machine (₹)", "type": "number", "min": 0},
        ],
    },
    {
        "key": "miscellaneous",
        "label": "9. Miscellaneous Charges",
        "fields": [
            {"name": "food_charges", "label": "Food Charges (patient/attendant) (₹)", "type": "number", "min": 0},
            {"name": "linen_charges", "label": "Linen Charges (₹)", "type": "number", "min": 0},
            {"name": "biomedical_waste_charges", "label": "Bio-medical Waste Charges (₹)", "type": "number", "min": 0},
            {"name": "wheelchair_stretcher_charges", "label": "Wheelchair / Stretcher Charges (₹)", "type": "number", "min": 0},
            {"name": "ambulance_charges", "label": "Ambulance Charges (₹)", "type": "number", "min": 0},
            {"name": "mortuary_services", "label": "Mortuary Services (₹)", "type": "number", "min": 0},
        ],
    },
    {
        "key": "packages",
        "label": "10. Package Billing",
        "fields": [
            {"name": "normal_delivery_package", "label": "Normal Delivery Package (₹)", "type": "number", "min": 0},
            {"name": "cesarean_section_package", "label": "Cesarean Section Package (₹)", "type": "number", "min": 0},
            {"name": "cataract_surgery_package", "label": "Cataract Surgery Package (₹)", "type": "number", "min": 0},
            {"name": "knee_replacement_package", "label": "Knee Replacement Package (₹)", "type": "number", "min": 0},
            {"name": "icu_package", "label": "ICU Package (₹)", "type": "number", "min": 0},
            {"name": "covid_treatment_package", "label": "COVID Treatment Package (₹)", "type": "number", "min": 0},
            {"name": "dialysis_package", "label": "Dialysis Package (₹)", "type": "number", "min": 0},
        ],
    },
]

FIELD_ORDER = [field["name"] for section in FIELD_SECTIONS for field in section["fields"]]
DOCTOR_FIELD_ORDER = [field["name"] for section in DOCTOR_FIELD_SECTIONS for field in section["fields"]]
OPD_FIELD_ORDER = [field["name"] for section in OPD_FIELD_SECTIONS for field in section["fields"]]
ADMISSION_FIELD_ORDER = [field["name"] for section in ADMISSION_FIELD_SECTIONS for field in section["fields"]]
CHARGE_FIELD_ORDER = [field["name"] for section in CHARGE_FIELD_SECTIONS for field in section["fields"]]
FIELD_MAP = {field["name"]: field for section in FIELD_SECTIONS for field in section["fields"]}
DOCTOR_FIELD_MAP = {field["name"]: field for section in DOCTOR_FIELD_SECTIONS for field in section["fields"]}
OPD_FIELD_MAP = {field["name"]: field for section in OPD_FIELD_SECTIONS for field in section["fields"]}
ADMISSION_FIELD_MAP = {field["name"]: field for section in ADMISSION_FIELD_SECTIONS for field in section["fields"]}
CHARGE_FIELD_MAP = {field["name"]: field for section in CHARGE_FIELD_SECTIONS for field in section["fields"]}
HEADERS = ["ID"] + [field_name for field_name in FIELD_ORDER]
DOCTOR_HEADERS = ["ID"] + [field_name for field_name in DOCTOR_FIELD_ORDER]
OPD_HEADERS = ["ID"] + [field_name for field_name in OPD_FIELD_ORDER]
ADMISSION_HEADERS = ["ID"] + [field_name for field_name in ADMISSION_FIELD_ORDER]
CHARGE_HEADERS = ["ID"] + [field_name for field_name in CHARGE_FIELD_ORDER]
REQUIRED_FIELDS = {field["name"] for section in FIELD_SECTIONS for field in section["fields"] if field.get("required")}
DOCTOR_REQUIRED_FIELDS = {field["name"] for section in DOCTOR_FIELD_SECTIONS for field in section["fields"] if field.get("required")}
OPD_REQUIRED_FIELDS = {field["name"] for section in OPD_FIELD_SECTIONS for field in section["fields"] if field.get("required")}
ADMISSION_REQUIRED_FIELDS = {field["name"] for section in ADMISSION_FIELD_SECTIONS for field in section["fields"] if field.get("required")}
CHARGE_REQUIRED_FIELDS = {field["name"] for section in CHARGE_FIELD_SECTIONS for field in section["fields"] if field.get("required")}

app = Flask(__name__)
app.config["SECRET_KEY"] = "dev-secret-key"

# Custom Jinja2 filter for safe admission ID formatting
@app.template_filter('format_admission_id')
def format_admission_id(value):
    """Safely format admission ID as ADM00003"""
    if value is None:
        return "N/A"
    try:
        if isinstance(value, str):
            value = int(float(value))
        else:
            value = int(value)
        return f"ADM{value:05d}"
    except (ValueError, TypeError):
        return str(value)


@dataclass
class Patient:
    patient_id: int
    full_name: str = ""
    gender: str = ""
    dob: str = ""
    age: str = ""
    blood_group: str = ""
    marital_status: str = ""
    photo_url: str = ""
    mobile_primary: str = ""
    mobile_alternate: str = ""
    email: str = ""
    address_permanent: str = ""
    address_local: str = ""
    aadhar_number: str = ""
    pan_number: str = ""
    hospital_id: str = ""
    govt_id_upload: str = ""
    emergency_name: str = ""
    emergency_relationship: str = ""
    emergency_mobile: str = ""
    emergency_address: str = ""
    allergies: str = ""
    existing_conditions: str = ""
    past_surgeries: str = ""
    current_medication: str = ""
    family_history: str = ""
    habits: str = ""
    diagnosis: str = ""
    treatment_for: str = ""
    procedures_planned: str = ""
    tests_recommended: str = ""
    billing_type: str = ""
    insurance_provider: str = ""
    policy_number: str = ""
    tpa_details: str = ""
    claim_type: str = ""
    treatment_consent: str = ""
    surgery_consent: str = ""
    privacy_consent: str = ""
    guardian_details: str = ""

    def to_row(self) -> List:
        return [self.patient_id] + [getattr(self, field_name, "") for field_name in FIELD_ORDER]


@dataclass
class Doctor:
    doctor_id: int
    full_name: str = ""
    gender: str = ""
    dob: str = ""
    age: str = ""
    contact_primary: str = ""
    contact_alternate: str = ""
    email: str = ""
    address: str = ""
    photo_url: str = ""
    qualification: str = ""
    specialization: str = ""
    sub_specialization: str = ""
    registration_number: str = ""
    experience_years: str = ""
    fellowships: str = ""
    doctor_type: str = ""
    department: str = ""
    joining_date: str = ""
    status: str = ""
    opd_fee_initial: str = ""
    opd_fee_followup: str = ""
    opd_days: str = ""
    opd_timings: str = ""
    opd_max_patients: str = ""
    ipd_visit_charge: str = ""
    icu_visit_charge: str = ""
    round_frequency: str = ""
    surgery_fee: str = ""
    ot_eligibility: str = ""
    surgery_types: str = ""
    weekly_schedule: str = ""
    locations: str = ""
    emergency_oncall: str = ""
    leaves_offdays: str = ""
    bank_account: str = ""
    ifsc_code: str = ""
    pan_number: str = ""
    upi_id: str = ""
    payment_mode: str = ""
    username: str = ""
    password: str = ""
    role: str = ""
    access_level: str = ""

    def to_row(self) -> List:
        return [self.doctor_id] + [getattr(self, field_name, "") for field_name in DOCTOR_FIELD_ORDER]


@dataclass
class OPD:
    opd_id: int
    patient_name: str = ""
    age: str = ""
    gender: str = ""
    mobile_number: str = ""
    opd_token: str = ""
    visit_date_time: str = ""
    department: str = ""
    doctor_name: str = ""
    visit_type: str = ""
    chief_complaint: str = ""
    diagnosis: str = ""
    treatment_for: str = ""
    consultation_fee: str = ""
    discount: str = ""
    payment_method: str = ""
    bill_number: str = ""
    address: str = ""
    allergies: str = ""
    existing_conditions: str = ""
    insurance: str = ""

    def to_row(self) -> List:
        return [self.opd_id] + [getattr(self, field_name, "") for field_name in OPD_FIELD_ORDER]


@dataclass
class Admission:
    admission_id: int
    admission_date_time: str = ""
    admission_type: str = ""
    visit_type: str = ""
    patient_id: str = ""
    patient_name: str = ""
    patient_age: str = ""
    patient_gender: str = ""
    attendant_name: str = ""
    attendant_relationship: str = ""
    attendant_mobile: str = ""
    attendant_address: str = ""
    chief_complaint: str = ""
    provisional_diagnosis: str = ""
    mode_of_arrival: str = ""
    referring_doctor: str = ""
    triage_category: str = ""
    admitting_department: str = ""
    admitting_consultant: str = ""
    assistant_doctor: str = ""
    clinical_unit: str = ""
    ward: str = ""
    room_type: str = ""
    room_number: str = ""
    bed_number: str = ""
    bed_status: str = ""
    expected_stay: str = ""
    billing_category: str = ""
    tariff_plan: str = ""
    initial_deposit: str = ""
    payment_mode: str = ""
    receipt_number: str = ""
    insurance_company: str = ""
    tpa_name: str = ""
    policy_number: str = ""
    policy_holder_name: str = ""
    policy_holder_relation: str = ""
    policy_validity_from: str = ""
    policy_validity_to: str = ""
    authorization_number: str = ""
    approved_amount: str = ""
    corporate_name: str = ""
    treatment_consent: str = ""
    icu_consent: str = ""
    surgery_consent: str = ""
    dnr_flag: str = ""
    consent_taken_from: str = ""
    consent_relative_name: str = ""
    consent_form_upload: str = ""
    created_by: str = ""
    created_date_time: str = ""
    last_updated_by: str = ""
    status: str = ""
    source_of_admission: str = ""
    current_ward: str = ""
    current_bed: str = ""
    transfer_history: str = ""
    clinical_severity: str = ""
    isolation_flag: str = ""
    discharge_date_time: str = ""
    discharge_type: str = ""
    final_diagnosis: str = ""
    condition_at_discharge: str = ""
    followup_date: str = ""
    followup_doctor: str = ""
    followup_department: str = ""

    def to_row(self) -> List:
        return [self.admission_id] + [getattr(self, field_name, "") for field_name in ADMISSION_FIELD_ORDER]


@dataclass
class ChargeMaster:
    charge_master_id: int
    # Registration & Administrative
    registration_fee: str = ""
    file_opening_charges: str = ""
    card_opd_slip_charges: str = ""
    admission_processing_fee: str = ""
    emergency_registration_fee: str = ""
    # Room / Bed Charges
    general_ward_bed: str = ""
    semi_private_room: str = ""
    private_room: str = ""
    deluxe_room: str = ""
    suite_room: str = ""
    icu: str = ""
    iccu: str = ""
    nicu_picu: str = ""
    ventilator_bed: str = ""
    isolation_room: str = ""
    # Nursing Charges
    nursing_care_charge: str = ""
    special_nursing_charge: str = ""
    attendant_charges: str = ""
    # Doctor Visit Charges
    opd_consultation_fee: str = ""
    opd_followup_fee: str = ""
    ipd_daily_visit_charge: str = ""
    icu_visit_charge: str = ""
    night_visit_charge: str = ""
    surgeon_visit_charge: str = ""
    # Procedures & Treatment
    dressing: str = ""
    nebulization: str = ""
    catheterization: str = ""
    injection_charges: str = ""
    iv_fluids_administration: str = ""
    enema: str = ""
    blood_transfusion: str = ""
    plaster_pop: str = ""
    wound_suturing: str = ""
    physiotherapy_session: str = ""
    dialysis_session: str = ""
    # OT / Procedure Room Charges
    ot_charges: str = ""
    minor_ot_charges: str = ""
    anesthesia_charges: str = ""
    anesthetist_visit_charge: str = ""
    surgeon_fee: str = ""
    assistant_surgeon_fee: str = ""
    recovery_room_charges: str = ""
    # Pharmacy / Medicine Charges
    tablets_charge: str = ""
    injections_charge: str = ""
    iv_fluids_charge: str = ""
    consumables_charge: str = ""
    surgical_consumables_charge: str = ""
    # Bedside Services
    oxygen_charges: str = ""
    ventilator_charges: str = ""
    defibrillator_usage: str = ""
    cpap_bipap_use: str = ""
    suction_machine: str = ""
    # Miscellaneous Charges
    food_charges: str = ""
    linen_charges: str = ""
    biomedical_waste_charges: str = ""
    wheelchair_stretcher_charges: str = ""
    ambulance_charges: str = ""
    mortuary_services: str = ""
    # Package Billing
    normal_delivery_package: str = ""
    cesarean_section_package: str = ""
    cataract_surgery_package: str = ""
    knee_replacement_package: str = ""
    icu_package: str = ""
    covid_treatment_package: str = ""
    dialysis_package: str = ""

    def to_row(self) -> List:
        return [self.charge_master_id] + [getattr(self, field_name, "") for field_name in CHARGE_FIELD_ORDER]


@dataclass
class Billing:
    bill_id: int
    bill_number: str = ""
    patient_id: str = ""
    patient_name: str = ""
    admission_id: str = ""
    billing_date: str = ""
    billing_type: str = ""  # OPD, IPD
    charges_json: str = ""  # JSON string of charges with quantities
    subtotal: str = ""
    discount: str = ""
    tax: str = ""
    total_amount: str = ""
    payment_status: str = ""  # Pending, Paid, Partial
    payment_mode: str = ""
    payment_reference: str = ""
    notes: str = ""
    bill_status: str = ""  # Draft, Final

    def to_row(self) -> List:
        return [
            self.bill_id,
            self.bill_number,
            self.patient_id,
            self.patient_name,
            self.admission_id,
            self.billing_date,
            self.billing_type,
            self.charges_json,
            self.subtotal,
            self.discount,
            self.tax,
            self.total_amount,
            self.payment_status,
            self.payment_mode,
            self.payment_reference,
            self.notes,
            self.bill_status,
        ]


BILLING_FIELD_ORDER = [
    "bill_number", "patient_id", "patient_name", "admission_id", "billing_date",
    "billing_type", "charges_json", "subtotal", "discount", "tax", "total_amount",
    "payment_status", "payment_mode", "payment_reference", "notes", "bill_status"
]


BILLING_HEADERS = ["ID"] + BILLING_FIELD_ORDER


@dataclass
class AdmissionCharge:
    charge_entry_id: int
    admission_id: str = ""
    patient_id: str = ""
    patient_name: str = ""
    billing_type: str = ""
    charges_json: str = ""
    subtotal: str = ""
    discount: str = ""
    tax: str = ""
    total_amount: str = ""
    status: str = "Pending"  # Pending, Merged
    created_at: str = ""

    def to_row(self) -> List:
        return [
            self.charge_entry_id,
            self.admission_id,
            self.patient_id,
            self.patient_name,
            self.billing_type,
            self.charges_json,
            self.subtotal,
            self.discount,
            self.tax,
            self.total_amount,
            self.status,
            self.created_at,
        ]


ADMISSION_CHARGE_FIELD_ORDER = [
    "admission_id",
    "patient_id",
    "patient_name",
    "billing_type",
    "charges_json",
    "subtotal",
    "discount",
    "tax",
    "total_amount",
    "status",
    "created_at",
]

ADMISSION_CHARGE_HEADERS = ["Charge Entry ID"] + ADMISSION_CHARGE_FIELD_ORDER


def _ensure_headers(ws):
    for idx, header in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=idx, value=header)


def _load_workbook():
    if not PATIENT_FILE.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        _ensure_headers(ws)
        wb.save(PATIENT_FILE)
        return wb
    wb = load_workbook(PATIENT_FILE)
    ws = wb[SHEET_NAME]
    _ensure_headers(ws)
    return wb


def _next_patient_id(ws) -> int:
    ids = [cell.value for cell in ws["A"][1:] if cell.value is not None]
    return (max(ids) + 1) if ids else 1


def _get_patients() -> List[Patient]:
    wb = _load_workbook()
    ws = wb[SHEET_NAME]
    patients: List[Patient] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        kwargs: Dict[str, str] = {"patient_id": int(row[0])}
        for idx, field_name in enumerate(FIELD_ORDER, start=1):
            kwargs[field_name] = str(row[idx]) if idx < len(row) and row[idx] is not None else ""
        patients.append(Patient(**kwargs))
    wb.close()
    return patients


def _find_patient(patient_id: int) -> Optional[Patient]:
    patients = _get_patients()
    for patient in patients:
        if patient.patient_id == patient_id:
            return patient
    return None


def _update_patient_row(patient: Patient):
    wb = _load_workbook()
    ws = wb[SHEET_NAME]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == patient.patient_id:
            for col_idx, field_name in enumerate(FIELD_ORDER, start=1):
                row[col_idx].value = getattr(patient, field_name, "")
            wb.save(PATIENT_FILE)
            wb.close()
            return
    wb.close()
    raise ValueError("Patient not found in workbook")


def _create_patient(payload: Dict[str, str]) -> Patient:
    wb = _load_workbook()
    ws = wb[SHEET_NAME]
    patient_id = _next_patient_id(ws)
    payload.setdefault("hospital_id", f"MRN{patient_id:05d}")
    patient = Patient(patient_id=patient_id, **payload)
    ws.append(patient.to_row())
    wb.save(PATIENT_FILE)
    wb.close()
    return patient


def _validate_payload(data: dict) -> Optional[str]:
    for field_name in REQUIRED_FIELDS:
        if not data.get(field_name, "").strip():
            return f"{field_name.replace('_', ' ').title()} is required."
    return None


def _calculate_age(date_str: str) -> str:
    try:
        dob = datetime.strptime(date_str, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return ""
    today = date.today()
    if dob > today:
        return ""
    years = today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day))
    return str(years)


def _extract_prefill(form_data, patient: Optional[Patient] = None) -> Dict[str, str]:
    result = {field_name: "" for field_name in FIELD_ORDER}
    if patient:
        for field_name in FIELD_ORDER:
            result[field_name] = getattr(patient, field_name, "")
        if patient.dob:
            result["age"] = _calculate_age(patient.dob) or patient.age
    for field_name in FIELD_ORDER:
        value = form_data.get(field_name)
        if value is not None:
            result[field_name] = value
    return result


def _coerce_checkbox(value: Optional[str]) -> str:
    return "Yes" if value in ("on", "yes", "Yes", "true", "True", "1") else "No"


def _extract_payload(form) -> Dict[str, str]:
    payload: Dict[str, str] = {}
    for field_name in FIELD_ORDER:
        field_config = FIELD_MAP[field_name]
        if field_config.get("auto"):
            continue
        if field_config.get("type") == "checkbox":
            payload[field_name] = _coerce_checkbox(form.get(field_name))
        else:
            payload[field_name] = form.get(field_name, "").strip()
    dob_value = payload.get("dob")
    if dob_value:
        calculated_age = _calculate_age(dob_value)
        if calculated_age:
            payload["age"] = calculated_age
    return payload


def _ensure_doctor_headers(ws):
    for idx, header in enumerate(DOCTOR_HEADERS, start=1):
        ws.cell(row=1, column=idx, value=header)


def _load_doctor_workbook():
    if not DOCTOR_FILE.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = DOCTOR_SHEET_NAME
        _ensure_doctor_headers(ws)
        wb.save(DOCTOR_FILE)
        return wb
    wb = load_workbook(DOCTOR_FILE)
    ws = wb[DOCTOR_SHEET_NAME]
    _ensure_doctor_headers(ws)
    return wb


def _next_doctor_id(ws) -> int:
    ids = [cell.value for cell in ws["A"][1:] if cell.value is not None]
    return (max(ids) + 1) if ids else 1


def _get_doctors() -> List[Doctor]:
    wb = _load_doctor_workbook()
    ws = wb[DOCTOR_SHEET_NAME]
    doctors: List[Doctor] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        kwargs: Dict[str, str] = {"doctor_id": int(row[0])}
        for idx, field_name in enumerate(DOCTOR_FIELD_ORDER, start=1):
            kwargs[field_name] = str(row[idx]) if idx < len(row) and row[idx] is not None else ""
        doctors.append(Doctor(**kwargs))
    wb.close()
    return doctors


def _find_doctor(doctor_id: int) -> Optional[Doctor]:
    doctors = _get_doctors()
    for doctor in doctors:
        if doctor.doctor_id == doctor_id:
            return doctor
    return None


def _update_doctor_row(doctor: Doctor):
    wb = _load_doctor_workbook()
    ws = wb[DOCTOR_SHEET_NAME]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == doctor.doctor_id:
            for col_idx, field_name in enumerate(DOCTOR_FIELD_ORDER, start=1):
                row[col_idx].value = getattr(doctor, field_name, "")
            wb.save(DOCTOR_FILE)
            wb.close()
            return
    wb.close()
    raise ValueError("Doctor not found in workbook")


def _create_doctor(payload: Dict[str, str]) -> Doctor:
    wb = _load_doctor_workbook()
    ws = wb[DOCTOR_SHEET_NAME]
    doctor_id = _next_doctor_id(ws)
    doctor = Doctor(doctor_id=doctor_id, **payload)
    ws.append(doctor.to_row())
    wb.save(DOCTOR_FILE)
    wb.close()
    return doctor


def _validate_doctor_payload(data: dict) -> Optional[str]:
    for field_name in DOCTOR_REQUIRED_FIELDS:
        if not data.get(field_name, "").strip():
            return f"{field_name.replace('_', ' ').title()} is required."
    return None


def _extract_doctor_prefill(form_data, doctor: Optional[Doctor] = None) -> Dict[str, str]:
    result = {field_name: "" for field_name in DOCTOR_FIELD_ORDER}
    if doctor:
        for field_name in DOCTOR_FIELD_ORDER:
            result[field_name] = getattr(doctor, field_name, "")
        if doctor.dob:
            result["age"] = _calculate_age(doctor.dob) or doctor.age
    for field_name in DOCTOR_FIELD_ORDER:
        value = form_data.get(field_name)
        if value is not None:
            result[field_name] = value
    return result


def _extract_doctor_payload(form) -> Dict[str, str]:
    payload: Dict[str, str] = {}
    for field_name in DOCTOR_FIELD_ORDER:
        field_config = DOCTOR_FIELD_MAP[field_name]
        if field_config.get("auto"):
            continue
        if field_config.get("type") == "checkbox":
            payload[field_name] = _coerce_checkbox(form.get(field_name))
        else:
            payload[field_name] = form.get(field_name, "").strip()
    dob_value = payload.get("dob")
    if dob_value:
        calculated_age = _calculate_age(dob_value)
        if calculated_age:
            payload["age"] = calculated_age
    return payload


def _ensure_opd_headers(ws):
    for idx, header in enumerate(OPD_HEADERS, start=1):
        ws.cell(row=1, column=idx, value=header)


def _load_opd_workbook():
    if not OPD_FILE.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = OPD_SHEET_NAME
        _ensure_opd_headers(ws)
        wb.save(OPD_FILE)
        return wb
    wb = load_workbook(OPD_FILE)
    ws = wb[OPD_SHEET_NAME]
    _ensure_opd_headers(ws)
    return wb


def _next_opd_id(ws) -> int:
    ids = [cell.value for cell in ws["A"][1:] if cell.value is not None]
    return (max(ids) + 1) if ids else 1


def _get_opd_records() -> List[OPD]:
    wb = _load_opd_workbook()
    ws = wb[OPD_SHEET_NAME]
    opd_records: List[OPD] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        try:
            # Handle both int and float from Excel
            opd_id_value = row[0]
            if isinstance(opd_id_value, float):
                opd_id_value = int(opd_id_value)
            elif isinstance(opd_id_value, str):
                opd_id_value = int(float(opd_id_value))
            else:
                opd_id_value = int(opd_id_value)
            kwargs: Dict[str, str] = {"opd_id": opd_id_value}
            for idx, field_name in enumerate(OPD_FIELD_ORDER, start=1):
                kwargs[field_name] = str(row[idx]) if idx < len(row) and row[idx] is not None else ""
            opd_records.append(OPD(**kwargs))
        except (ValueError, TypeError) as e:
            # Skip invalid rows
            continue
    wb.close()
    return opd_records


def _find_opd(opd_id: int) -> Optional[OPD]:
    opd_records = _get_opd_records()
    # Ensure opd_id is an int for comparison
    search_id = int(opd_id)
    for opd in opd_records:
        # Ensure both are ints for comparison
        if int(opd.opd_id) == search_id:
            return opd
    return None


def _update_opd_row(opd: OPD):
    wb = _load_opd_workbook()
    ws = wb[OPD_SHEET_NAME]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == opd.opd_id:
            for col_idx, field_name in enumerate(OPD_FIELD_ORDER, start=1):
                row[col_idx].value = getattr(opd, field_name, "")
            wb.save(OPD_FILE)
            wb.close()
            return
    wb.close()
    raise ValueError("OPD record not found in workbook")


def _create_opd(payload: Dict[str, str]) -> OPD:
    wb = _load_opd_workbook()
    ws = wb[OPD_SHEET_NAME]
    opd_id = _next_opd_id(ws)
    payload.setdefault("opd_token", f"OPD{opd_id:05d}")
    payload.setdefault("bill_number", f"BILL{opd_id:05d}")
    if not payload.get("visit_date_time"):
        payload["visit_date_time"] = datetime.now().strftime("%Y-%m-%dT%H:%M")
    opd = OPD(opd_id=opd_id, **payload)
    ws.append(opd.to_row())
    wb.save(OPD_FILE)
    wb.close()
    return opd


def _validate_opd_payload(data: dict) -> Optional[str]:
    for field_name in OPD_REQUIRED_FIELDS:
        if not data.get(field_name, "").strip():
            return f"{field_name.replace('_', ' ').title()} is required."
    return None


def _extract_opd_prefill(form_data, opd: Optional[OPD] = None) -> Dict[str, str]:
    result = {field_name: "" for field_name in OPD_FIELD_ORDER}
    if opd:
        for field_name in OPD_FIELD_ORDER:
            result[field_name] = getattr(opd, field_name, "")
    for field_name in OPD_FIELD_ORDER:
        value = form_data.get(field_name)
        if value is not None:
            result[field_name] = value
    if not result.get("visit_date_time") and not opd:
        result["visit_date_time"] = datetime.now().strftime("%Y-%m-%dT%H:%M")
    return result


def _extract_opd_payload(form) -> Dict[str, str]:
    payload: Dict[str, str] = {}
    for field_name in OPD_FIELD_ORDER:
        field_config = OPD_FIELD_MAP[field_name]
        if field_config.get("auto"):
            continue
        if field_config.get("type") == "checkbox":
            payload[field_name] = _coerce_checkbox(form.get(field_name))
        else:
            payload[field_name] = form.get(field_name, "").strip()
    if not payload.get("visit_date_time"):
        payload["visit_date_time"] = datetime.now().strftime("%Y-%m-%dT%H:%M")
    return payload


def _ensure_admission_headers(ws):
    for idx, header in enumerate(ADMISSION_HEADERS, start=1):
        ws.cell(row=1, column=idx, value=header)


def _load_admission_workbook():
    if not ADMISSION_FILE.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = ADMISSION_SHEET_NAME
        _ensure_admission_headers(ws)
        wb.save(ADMISSION_FILE)
        return wb
    wb = load_workbook(ADMISSION_FILE)
    ws = wb[ADMISSION_SHEET_NAME]
    _ensure_admission_headers(ws)
    return wb


def _next_admission_id(ws) -> int:
    ids = [cell.value for cell in ws["A"][1:] if cell.value is not None]
    return (max(ids) + 1) if ids else 1


def _get_admissions() -> List[Admission]:
    wb = _load_admission_workbook()
    ws = wb[ADMISSION_SHEET_NAME]
    admissions: List[Admission] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        try:
            admission_id_value = row[0]
            if isinstance(admission_id_value, float):
                admission_id_value = int(admission_id_value)
            elif isinstance(admission_id_value, str):
                admission_id_value = int(float(admission_id_value))
            else:
                admission_id_value = int(admission_id_value)
            kwargs: Dict[str, Any] = {"admission_id": int(admission_id_value)}
            # Get all field names from Admission dataclass
            admission_fields = {f.name for f in Admission.__dataclass_fields__.values()}
            for idx, field_name in enumerate(ADMISSION_FIELD_ORDER, start=1):
                # Only include fields that exist in the Admission dataclass
                if field_name in admission_fields:
                    kwargs[field_name] = str(row[idx]) if idx < len(row) and row[idx] is not None else ""
            admissions.append(Admission(**kwargs))
        except (ValueError, TypeError):
            continue
    wb.close()
    return admissions


def _find_admission(admission_id: int) -> Optional[Admission]:
    admissions = _get_admissions()
    search_id = int(admission_id)
    for admission in admissions:
        if int(admission.admission_id) == search_id:
            return admission
    return None


def _update_admission_row(admission: Admission):
    wb = _load_admission_workbook()
    ws = wb[ADMISSION_SHEET_NAME]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == admission.admission_id:
            for col_idx, field_name in enumerate(ADMISSION_FIELD_ORDER, start=1):
                row[col_idx].value = getattr(admission, field_name, "")
            wb.save(ADMISSION_FILE)
            wb.close()
            return
    wb.close()
    raise ValueError("Admission not found in workbook")


def _create_admission(payload: Dict[str, str]) -> Admission:
    wb = _load_admission_workbook()
    ws = wb[ADMISSION_SHEET_NAME]
    admission_id = _next_admission_id(ws)
    if not payload.get("admission_date_time"):
        payload["admission_date_time"] = datetime.now().strftime("%Y-%m-%dT%H:%M")
    if not payload.get("created_date_time"):
        payload["created_date_time"] = datetime.now().strftime("%Y-%m-%dT%H:%M")
    admission = Admission(admission_id=admission_id, **payload)
    ws.append(admission.to_row())
    wb.save(ADMISSION_FILE)
    wb.close()
    return admission


def _validate_admission_payload(data: dict) -> Optional[str]:
    for field_name in ADMISSION_REQUIRED_FIELDS:
        if not data.get(field_name, "").strip():
            return f"{field_name.replace('_', ' ').title()} is required."
    return None


def _extract_admission_prefill(form_data, admission: Optional[Admission] = None) -> Dict[str, str]:
    result = {field_name: "" for field_name in ADMISSION_FIELD_ORDER}
    if admission:
        for field_name in ADMISSION_FIELD_ORDER:
            result[field_name] = getattr(admission, field_name, "")
    for field_name in ADMISSION_FIELD_ORDER:
        value = form_data.get(field_name)
        if value is not None:
            result[field_name] = value
    if not result.get("admission_date_time") and not admission:
        result["admission_date_time"] = datetime.now().strftime("%Y-%m-%dT%H:%M")
    if not result.get("created_date_time") and not admission:
        result["created_date_time"] = datetime.now().strftime("%Y-%m-%dT%H:%M")
    return result


def _extract_admission_payload(form) -> Dict[str, str]:
    payload: Dict[str, str] = {}
    for field_name in ADMISSION_FIELD_ORDER:
        field_config = ADMISSION_FIELD_MAP[field_name]
        if field_config.get("auto"):
            continue
        if field_config.get("type") == "checkbox":
            payload[field_name] = _coerce_checkbox(form.get(field_name))
        else:
            payload[field_name] = form.get(field_name, "").strip()
    if not payload.get("admission_date_time"):
        payload["admission_date_time"] = datetime.now().strftime("%Y-%m-%dT%H:%M")
    if not payload.get("created_date_time"):
        payload["created_date_time"] = datetime.now().strftime("%Y-%m-%dT%H:%M")
    return payload


def _load_charge_workbook():
    if not CHARGE_FILE.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = CHARGE_SHEET_NAME
        for idx, header in enumerate(CHARGE_HEADERS, start=1):
            ws.cell(row=1, column=idx, value=header)
        # Create the single charge master record with ID=1
        charge_master = ChargeMaster(charge_master_id=1)
        ws.append(charge_master.to_row())
        wb.save(CHARGE_FILE)
        wb.close()
    return load_workbook(CHARGE_FILE)


def _get_charge_master() -> Optional[ChargeMaster]:
    """Get the single charge master record (always ID=1)"""
    wb = _load_charge_workbook()
    ws = wb[CHARGE_SHEET_NAME]
    # Look for row with ID=1
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        try:
            charge_id_value = row[0]
            if isinstance(charge_id_value, float):
                charge_id_value = int(charge_id_value)
            elif isinstance(charge_id_value, str):
                charge_id_value = int(float(charge_id_value))
            else:
                charge_id_value = int(charge_id_value)
            
            if charge_id_value == 1:
                kwargs: Dict[str, Any] = {"charge_master_id": 1}
                for idx, field_name in enumerate(CHARGE_FIELD_ORDER, start=1):
                    kwargs[field_name] = str(row[idx]) if idx < len(row) and row[idx] is not None else ""
                wb.close()
                return ChargeMaster(**kwargs)
        except (ValueError, TypeError):
            continue
    wb.close()
    # If no record exists, create one
    return ChargeMaster(charge_master_id=1)


def _update_charge_master(charge_master: ChargeMaster):
    """Update the single charge master record"""
    wb = _load_charge_workbook()
    ws = wb[CHARGE_SHEET_NAME]
    # Find row with ID=1
    for row in ws.iter_rows(min_row=2):
        if row[0].value:
            try:
                row_id = int(float(str(row[0].value)))
                if row_id == 1:
                    for col_idx, field_name in enumerate(CHARGE_FIELD_ORDER, start=1):
                        row[col_idx].value = getattr(charge_master, field_name, "")
                    wb.save(CHARGE_FILE)
                    wb.close()
                    return
            except (ValueError, TypeError):
                continue
    # If no row found, append it
    ws.append(charge_master.to_row())
    wb.save(CHARGE_FILE)
    wb.close()


def _save_charge_master(payload: Dict[str, str]) -> ChargeMaster:
    """Save/update the charge master record"""
    charge_master = _get_charge_master()
    if not charge_master:
        charge_master = ChargeMaster(charge_master_id=1)
    
    # Update all fields from payload
    for field_name in CHARGE_FIELD_ORDER:
        if field_name in payload:
            setattr(charge_master, field_name, payload[field_name])
    
    _update_charge_master(charge_master)
    return charge_master


def _prepare_charge_form_data(charge_master: Optional[ChargeMaster] = None, form_data: Optional[Dict] = None) -> Dict[str, str]:
    if form_data is None:
        form_data = {}
    if charge_master is None:
        charge_master = _get_charge_master()
    
    result = {field_name: "" for field_name in CHARGE_FIELD_ORDER}
    if charge_master:
        for field_name in CHARGE_FIELD_ORDER:
            result[field_name] = getattr(charge_master, field_name, "")
    for field_name in CHARGE_FIELD_ORDER:
        value = form_data.get(field_name)
        if value is not None:
            result[field_name] = value
    return result


def _extract_charge_payload(form) -> Dict[str, str]:
    payload: Dict[str, str] = {}
    for field_name in CHARGE_FIELD_ORDER:
        field_config = CHARGE_FIELD_MAP[field_name]
        if field_config.get("auto"):
            continue
        if field_config.get("type") == "checkbox":
            payload[field_name] = _coerce_checkbox(form.get(field_name))
        else:
            value = form.get(field_name, "").strip()
            payload[field_name] = value
    return payload


def _load_billing_workbook():
    if not BILLING_FILE.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = BILLING_SHEET_NAME
        for idx, header in enumerate(BILLING_HEADERS, start=1):
            ws.cell(row=1, column=idx, value=header)
        wb.save(BILLING_FILE)
        wb.close()
    return load_workbook(BILLING_FILE)


def _get_bills() -> List[Billing]:
    wb = _load_billing_workbook()
    ws = wb[BILLING_SHEET_NAME]
    bills: List[Billing] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        try:
            bill_id_value = row[0]
            if isinstance(bill_id_value, float):
                bill_id_value = int(bill_id_value)
            elif isinstance(bill_id_value, str):
                bill_id_value = int(float(bill_id_value))
            else:
                bill_id_value = int(bill_id_value)
            
            kwargs: Dict[str, Any] = {"bill_id": bill_id_value}
            for idx, field_name in enumerate(BILLING_FIELD_ORDER, start=1):
                kwargs[field_name] = str(row[idx]) if idx < len(row) and row[idx] is not None else ""
            bills.append(Billing(**kwargs))
        except (ValueError, TypeError):
            continue
    wb.close()
    return bills


def _find_bill(bill_id: int) -> Optional[Billing]:
    all_bills = _get_bills()
    for bill in all_bills:
        if bill.bill_id == bill_id:
            return bill
    return None


def _next_bill_id(ws) -> int:
    ids = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0] is not None:
            try:
                ids.append(int(float(str(row[0]))))
            except (ValueError, TypeError):
                continue
    return (max(ids) + 1) if ids else 1


def _create_bill(payload: Dict[str, str]) -> Billing:
    wb = _load_billing_workbook()
    ws = wb[BILLING_SHEET_NAME]
    bill_id = _next_bill_id(ws)
    
    # Generate bill number if not provided
    if not payload.get("bill_number"):
        payload["bill_number"] = f"BILL{bill_id:06d}"
    
    if not payload.get("billing_date"):
        payload["billing_date"] = datetime.now().strftime("%Y-%m-%d")
    
    kwargs: Dict[str, Any] = {"bill_id": bill_id}
    for field_name in BILLING_FIELD_ORDER:
        kwargs[field_name] = payload.get(field_name, "")
    
    new_bill = Billing(**kwargs)
    ws.append(new_bill.to_row())
    wb.save(BILLING_FILE)
    wb.close()
    return new_bill


def _update_bill_row(bill: Billing):
    wb = _load_billing_workbook()
    ws = wb[BILLING_SHEET_NAME]
    for row in ws.iter_rows(min_row=2):
        if row[0].value:
            try:
                row_id = int(float(str(row[0].value)))
                if row_id == bill.bill_id:
                    for col_idx, field_name in enumerate(BILLING_FIELD_ORDER, start=1):
                        row[col_idx].value = getattr(bill, field_name, "")
                    wb.save(BILLING_FILE)
                    wb.close()
                    return
            except (ValueError, TypeError):
                continue
    wb.close()


def _load_admission_charges_workbook():
    if not ADMISSION_CHARGES_FILE.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = ADMISSION_CHARGES_SHEET_NAME
        for idx, header in enumerate(ADMISSION_CHARGE_HEADERS, start=1):
            ws.cell(row=1, column=idx, value=header)
        wb.save(ADMISSION_CHARGES_FILE)
        wb.close()
    return load_workbook(ADMISSION_CHARGES_FILE)


def _get_admission_charges() -> List[AdmissionCharge]:
    wb = _load_admission_charges_workbook()
    ws = wb[ADMISSION_CHARGES_SHEET_NAME]
    entries: List[AdmissionCharge] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        try:
            entry_id_value = row[0]
            if isinstance(entry_id_value, float):
                entry_id_value = int(entry_id_value)
            elif isinstance(entry_id_value, str):
                entry_id_value = int(float(entry_id_value))
            else:
                entry_id_value = int(entry_id_value)
            kwargs: Dict[str, Any] = {"charge_entry_id": entry_id_value}
            for idx, field_name in enumerate(ADMISSION_CHARGE_FIELD_ORDER, start=1):
                kwargs[field_name] = str(row[idx]) if idx < len(row) and row[idx] is not None else ""
            entries.append(AdmissionCharge(**kwargs))
        except (ValueError, TypeError):
            continue
    wb.close()
    return entries


def _next_admission_charge_id(ws) -> int:
    ids = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0] is not None:
            try:
                ids.append(int(float(str(row[0]))))
            except (ValueError, TypeError):
                continue
    return (max(ids) + 1) if ids else 1


def _create_admission_charge(payload: Dict[str, str]) -> AdmissionCharge:
    wb = _load_admission_charges_workbook()
    ws = wb[ADMISSION_CHARGES_SHEET_NAME]
    entry_id = _next_admission_charge_id(ws)

    if not payload.get("created_at"):
        payload["created_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    kwargs: Dict[str, Any] = {"charge_entry_id": entry_id}
    for field_name in ADMISSION_CHARGE_FIELD_ORDER:
        kwargs[field_name] = payload.get(field_name, "")

    new_entry = AdmissionCharge(**kwargs)
    ws.append(new_entry.to_row())
    wb.save(ADMISSION_CHARGES_FILE)
    wb.close()
    return new_entry


def _update_admission_charge_row(entry: AdmissionCharge):
    wb = _load_admission_charges_workbook()
    ws = wb[ADMISSION_CHARGES_SHEET_NAME]
    for row in ws.iter_rows(min_row=2):
        if row[0].value:
            try:
                row_id = int(float(str(row[0].value)))
                if row_id == entry.charge_entry_id:
                    for col_idx, field_name in enumerate(ADMISSION_CHARGE_FIELD_ORDER, start=1):
                        row[col_idx].value = getattr(entry, field_name, "")
                    wb.save(ADMISSION_CHARGES_FILE)
                    wb.close()
                    return
            except (ValueError, TypeError):
                continue
    wb.close()


def _collect_admission_billing_state(admission_id: str):
    """Gather charge history, usage trackers, and bills for a specific admission."""
    charge_entries: List[AdmissionCharge] = []
    pending_charge_entries: List[AdmissionCharge] = []
    used_registration_charges = set()
    used_room_bed_charges = set()
    total_nursing_care_days = 0

    registration_charges = {
        "registration_fee",
        "file_opening_charges",
        "card_opd_slip_charges",
        "admission_processing_fee",
        "emergency_registration_fee",
    }
    room_bed_charges_set = {
        "general_ward_bed",
        "semi_private_room",
        "private_room",
        "deluxe_room",
        "suite_room",
        "icu",
        "iccu",
        "nicu_picu",
        "ventilator_bed",
        "isolation_room",
    }

    def _update_tracking(charges_list: List[Dict[str, Any]]):
        nonlocal total_nursing_care_days
        for charge in charges_list:
            charge_code = charge.get("charge_code", "")
            try:
                quantity = float(charge.get("quantity", 0))
            except (ValueError, TypeError):
                quantity = 0
            if charge_code in registration_charges and quantity > 0:
                used_registration_charges.add(charge_code)
            if charge_code in room_bed_charges_set and quantity > 0:
                used_room_bed_charges.add(charge_code)
            if charge_code == "nursing_care_charge" and quantity > 0:
                total_nursing_care_days += quantity

    # Collect admission charge entries
    for entry in _get_admission_charges():
        if entry.admission_id and str(entry.admission_id) == str(admission_id):
            parsed_charges: List[Dict[str, Any]] = []
            if entry.charges_json:
                try:
                    parsed_charges = json.loads(entry.charges_json)
                except (json.JSONDecodeError, TypeError):
                    parsed_charges = []
            entry.parsed_charges = parsed_charges
            charge_entries.append(entry)
            if entry.status == "Pending":
                pending_charge_entries.append(entry)
            _update_tracking(parsed_charges)

    existing_bills: List[Billing] = []
    for bill in _get_bills():
        if bill.admission_id and str(bill.admission_id) == str(admission_id):
            if bill.bill_status == "Merged":
                continue
            parsed_charges: List[Dict[str, Any]] = []
            if bill.charges_json:
                try:
                    parsed_charges = json.loads(bill.charges_json)
                except (json.JSONDecodeError, TypeError):
                    parsed_charges = []
            bill.parsed_charges = parsed_charges
            existing_bills.append(bill)
            if bill.bill_status in ["Draft", "Final"]:
                _update_tracking(parsed_charges)

    existing_bills.sort(key=lambda x: x.billing_date or "", reverse=True)
    charge_entries.sort(key=lambda x: x.created_at or "", reverse=True)

    return {
        "charge_entries": charge_entries,
        "pending_charge_entries": pending_charge_entries,
        "existing_bills": existing_bills,
        "used_registration_charges": used_registration_charges,
        "used_room_bed_charges": used_room_bed_charges,
        "total_nursing_care_days": total_nursing_care_days,
    }




@app.route("/", methods=["GET"])
def index():
    all_patients = _get_patients()
    search_query = request.args.get("search", "").strip()
    if search_query:
        lowered = search_query.lower()
        search_results = [
            patient
            for patient in all_patients
            if lowered in str(patient.patient_id).lower()
            or lowered in patient.full_name.lower()
            or lowered in patient.hospital_id.lower()
            or lowered in patient.mobile_primary.lower()
        ]
    else:
        search_results = []

    selected_patient = None
    selected_id = request.args.get("selected_id", "").strip()
    if selected_id.isdigit():
        selected_patient = _find_patient(int(selected_id))

    form_data = _extract_prefill({}, selected_patient) if selected_patient else _extract_prefill({})

    return render_template(
        "index.html",
        sections=FIELD_SECTIONS,
        form_data=form_data,
        error=None,
        success=request.args.get("success"),
        search_query=search_query,
        search_results=search_results,
        selected_patient=selected_patient,
    )


@app.post("/patients")
def create_patient():
    form = request.form
    error = _validate_payload(form)
    if error:
        all_patients = _get_patients()
        return (
            render_template(
                "index.html",
                sections=FIELD_SECTIONS,
                form_data=_extract_prefill(form),
                error=error,
                success=None,
                search_query="",
                search_results=all_patients[:5],
                selected_patient=None,
            ),
            400,
        )
    payload = _extract_payload(form)
    patient = _create_patient(payload)
    name = patient.full_name or "Unnamed Patient"
    mobile = patient.mobile_primary or "N/A"
    success_msg = f"Patient created: {name} | Mobile: {mobile}"
    return redirect(url_for("index", success=success_msg))


@app.route("/patients/<int:patient_id>/edit", methods=["GET", "POST"])
def edit_patient(patient_id: int):
    patient = _find_patient(patient_id)
    if not patient:
        abort(404)

    if request.method == "POST":
        form = request.form
        error = _validate_payload(form)
        if error:
            return (
                render_template(
                    "edit.html",
                    patient=patient,
                    sections=FIELD_SECTIONS,
                    form_data=_extract_prefill(form, patient),
                    error=error,
                    success=None,
                ),
                400,
            )
        payload = _extract_payload(form)
        payload["hospital_id"] = patient.hospital_id
        updated_patient = Patient(patient_id=patient_id, **payload)
        _update_patient_row(updated_patient)
        return redirect(url_for("index", success="Patient updated successfully."))

    return render_template(
        "edit.html",
        patient=patient,
        sections=FIELD_SECTIONS,
        form_data=_extract_prefill({}, patient),
        error=None,
        success=None,
    )


@app.route("/doctors", methods=["GET"])
def index_doctors():
    all_doctors = _get_doctors()
    search_query = request.args.get("search", "").strip()
    if search_query:
        lowered = search_query.lower()
        search_results = [
            doctor
            for doctor in all_doctors
            if lowered in str(doctor.doctor_id).lower()
            or lowered in doctor.full_name.lower()
            or lowered in doctor.registration_number.lower()
            or lowered in doctor.contact_primary.lower()
        ]
    else:
        search_results = []

    selected_doctor = None
    selected_id = request.args.get("selected_id", "").strip()
    if selected_id.isdigit():
        selected_doctor = _find_doctor(int(selected_id))

    form_data = _extract_doctor_prefill({}, selected_doctor) if selected_doctor else _extract_doctor_prefill({})

    return render_template(
        "index_doctor.html",
        sections=DOCTOR_FIELD_SECTIONS,
        form_data=form_data,
        error=None,
        success=request.args.get("success"),
        search_query=search_query,
        search_results=search_results,
        selected_doctor=selected_doctor,
    )


@app.post("/doctors")
def create_doctor():
    form = request.form
    error = _validate_doctor_payload(form)
    if error:
        all_doctors = _get_doctors()
        return (
            render_template(
                "index_doctor.html",
                sections=DOCTOR_FIELD_SECTIONS,
                form_data=_extract_doctor_prefill(form),
                error=error,
                success=None,
                search_query="",
                search_results=all_doctors[:5],
                selected_doctor=None,
            ),
            400,
        )
    payload = _extract_doctor_payload(form)
    doctor = _create_doctor(payload)
    name = doctor.full_name or "Unnamed Doctor"
    mobile = doctor.contact_primary or "N/A"
    qualification = doctor.qualification or "N/A"
    success_msg = f"Doctor created: {name} | Mobile: {mobile} | Qualification: {qualification}"
    return redirect(url_for("index_doctors", success=success_msg))


@app.route("/doctors/<int:doctor_id>/edit", methods=["GET", "POST"])
def edit_doctor(doctor_id: int):
    doctor = _find_doctor(doctor_id)
    if not doctor:
        abort(404)

    if request.method == "POST":
        form = request.form
        error = _validate_doctor_payload(form)
        if error:
            return (
                render_template(
                    "edit_doctor.html",
                    doctor=doctor,
                    sections=DOCTOR_FIELD_SECTIONS,
                    form_data=_extract_doctor_prefill(form, doctor),
                    error=error,
                    success=None,
                ),
                400,
            )
        payload = _extract_doctor_payload(form)
        updated_doctor = Doctor(doctor_id=doctor_id, **payload)
        _update_doctor_row(updated_doctor)
        return redirect(url_for("index_doctors", success="Doctor updated successfully."))

    return render_template(
        "edit_doctor.html",
        doctor=doctor,
        sections=DOCTOR_FIELD_SECTIONS,
        form_data=_extract_doctor_prefill({}, doctor),
        error=None,
        success=None,
    )


@app.route("/patients/view-all", methods=["GET"])
def view_all_patients():
    all_patients = _get_patients()
    all_opd = _get_opd_records()
    search_query = request.args.get("search", "").strip()
    
    # Convert OPD records to a unified format for display
    combined_records = []
    
    # Add regular patients
    for patient in all_patients:
        combined_records.append({
            "type": "patient",
            "id": patient.patient_id,
            "name": patient.full_name,
            "mobile": patient.mobile_primary,
            "age": patient.age,
            "gender": patient.gender,
            "mrn": patient.hospital_id,
            "email": patient.email,
            "aadhar": patient.aadhar_number,
            "record": patient,
        })
    
    # Add OPD patients
    for opd in all_opd:
        combined_records.append({
            "type": "opd",
            "id": opd.opd_id,
            "name": opd.patient_name,
            "mobile": opd.mobile_number,
            "age": opd.age,
            "gender": opd.gender,
            "mrn": opd.opd_token,
            "email": "",
            "aadhar": "",
            "record": opd,
        })
    
    # Filter combined records if search query exists
    if search_query:
        lowered = search_query.lower()
        filtered_records = [
            record
            for record in combined_records
            if lowered in str(record["id"]).lower()
            or lowered in (record["name"] or "").lower()
            or lowered in (record["mrn"] or "").lower()
            or lowered in (record["mobile"] or "").lower()
            or lowered in (record["email"] or "").lower()
            or lowered in (record["aadhar"] or "").lower()
        ]
    else:
        filtered_records = combined_records
    
    page = request.args.get("page", "1")
    try:
        page = int(page)
        if page < 1:
            page = 1
    except (ValueError, TypeError):
        page = 1
    
    per_page = 20
    total_records = len(filtered_records)
    total_pages = (total_records + per_page - 1) // per_page if total_records > 0 else 1
    
    if page > total_pages:
        page = total_pages
    
    start_idx = (page - 1) * per_page
    end_idx = start_idx + per_page
    records_page = filtered_records[start_idx:end_idx]
    
    return render_template(
        "view_all_patients.html",
        records=records_page,
        page=page,
        total_pages=total_pages,
        total_patients=total_records,
        per_page=per_page,
        search_query=search_query,
    )


@app.route("/doctors/view-all", methods=["GET"])
def view_all_doctors():
    all_doctors = _get_doctors()
    search_query = request.args.get("search", "").strip()
    
    # Filter doctors if search query exists
    if search_query:
        lowered = search_query.lower()
        filtered_doctors = [
            doctor
            for doctor in all_doctors
            if lowered in str(doctor.doctor_id).lower()
            or lowered in doctor.full_name.lower()
            or lowered in doctor.registration_number.lower()
            or lowered in doctor.contact_primary.lower()
            or lowered in (doctor.doctor_type or "").lower()
            or lowered in (doctor.department or "").lower()
            or lowered in (doctor.specialization or "").lower()
        ]
    else:
        filtered_doctors = all_doctors
    
    page = request.args.get("page", "1")
    try:
        page = int(page)
        if page < 1:
            page = 1
    except (ValueError, TypeError):
        page = 1
    
    per_page = 20
    total_doctors = len(filtered_doctors)
    total_pages = (total_doctors + per_page - 1) // per_page if total_doctors > 0 else 1
    
    if page > total_pages:
        page = total_pages
    
    start_idx = (page - 1) * per_page
    end_idx = start_idx + per_page
    doctors_page = filtered_doctors[start_idx:end_idx]
    
    return render_template(
        "view_all_doctors.html",
        doctors=doctors_page,
        page=page,
        total_pages=total_pages,
        total_doctors=total_doctors,
        per_page=per_page,
        search_query=search_query,
    )


@app.route("/api/patients/search", methods=["GET"])
def search_patients_api():
    query = request.args.get("q", "").strip().lower()
    if not query:
        return jsonify({"patients": []})
    
    results = []
    
    # Search in Patient table
    all_patients = _get_patients()
    for patient in all_patients:
        if (query in patient.full_name.lower()
            or query in (patient.hospital_id or "").lower()
            or query in (patient.mobile_primary or "").lower()
            or query in str(patient.patient_id).lower()):
            results.append({
                "id": patient.patient_id,
                "name": patient.full_name,
                "age": patient.age or "",
                "gender": patient.gender or "",
                "mobile": patient.mobile_primary or "",
                "mrn": patient.hospital_id or "",
                "type": "patient",
            })
    
    # Search in OPD table - search by patient name only
    all_opd = _get_opd_records()
    for opd in all_opd:
        if query in (opd.patient_name or "").lower():
            results.append({
                "id": opd.opd_id,
                "name": opd.patient_name or "",
                "age": opd.age or "",
                "gender": opd.gender or "",
                "mobile": opd.mobile_number or "",
                "mrn": opd.opd_token or f"OPD{opd.opd_id:05d}",
                "type": "opd",
            })
    
    # Limit to 20 results
    return jsonify({"patients": results[:20]})


@app.route("/api/doctors/search", methods=["GET"])
def search_doctors_api():
    query = request.args.get("q", "").strip().lower()
    if not query:
        return jsonify({"doctors": []})
    
    all_doctors = _get_doctors()
    results = [
        {
            "id": doctor.doctor_id,
            "name": doctor.full_name,
            "specialization": doctor.specialization or "",
            "department": doctor.department or "",
            "registration": doctor.registration_number or "",
        }
        for doctor in all_doctors
        if query in doctor.full_name.lower()
        or query in (doctor.specialization or "").lower()
        or query in (doctor.department or "").lower()
        or query in str(doctor.doctor_id).lower()
        or query in (doctor.registration_number or "").lower()
    ]
    return jsonify({"doctors": results[:20]})  # Limit to 20 results


@app.route("/opd", methods=["GET"])
def index_opd():
    all_opd = _get_opd_records()
    search_query = request.args.get("search", "").strip()
    if search_query:
        lowered = search_query.lower()
        search_results = [
            opd
            for opd in all_opd
            if lowered in str(opd.opd_id).lower()
            or lowered in opd.patient_name.lower()
            or lowered in opd.mobile_number.lower()
            or lowered in opd.opd_token.lower()
            or lowered in opd.bill_number.lower()
        ]
    else:
        search_results = []

    selected_opd = None
    selected_id = request.args.get("selected_id", "").strip()
    if selected_id:
        try:
            opd_id = int(selected_id)
            selected_opd = _find_opd(opd_id)
            # If not found by ID, try to find by matching in all records (fallback)
            if not selected_opd:
                for opd in all_opd:
                    if str(opd.opd_id) == selected_id or str(opd.opd_id) == str(opd_id):
                        selected_opd = opd
                        break
        except (ValueError, TypeError):
            selected_opd = None

    form_data = _extract_opd_prefill({}, selected_opd) if selected_opd else _extract_opd_prefill({})

    return render_template(
        "index_opd.html",
        sections=OPD_FIELD_SECTIONS,
        form_data=form_data,
        error=None,
        success=request.args.get("success"),
        search_query=search_query,
        search_results=search_results,
        selected_opd=selected_opd,
    )


@app.post("/opd")
def create_opd():
    form = request.form
    error = _validate_opd_payload(form)
    if error:
        all_opd = _get_opd_records()
        return (
            render_template(
                "index_opd.html",
                sections=OPD_FIELD_SECTIONS,
                form_data=_extract_opd_prefill(form),
                error=error,
                success=None,
                search_query="",
                search_results=all_opd[:5],
                selected_opd=None,
            ),
            400,
        )
    payload = _extract_opd_payload(form)
    opd = _create_opd(payload)
    name = opd.patient_name or "Unnamed Patient"
    mobile = opd.mobile_number or "N/A"
    token = opd.opd_token or "N/A"
    success_msg = f"OPD record created: {name} | Mobile: {mobile} | Token: {token}"
    return redirect(url_for("index_opd", success=success_msg))


@app.route("/opd/<int:opd_id>/edit", methods=["GET", "POST"])
def edit_opd(opd_id: int):
    opd = _find_opd(opd_id)
    if not opd:
        abort(404)

    if request.method == "POST":
        form = request.form
        error = _validate_opd_payload(form)
        if error:
            return (
                render_template(
                    "edit_opd.html",
                    opd=opd,
                    sections=OPD_FIELD_SECTIONS,
                    form_data=_extract_opd_prefill(form, opd),
                    error=error,
                    success=None,
                ),
                400,
            )
        payload = _extract_opd_payload(form)
        payload["opd_token"] = opd.opd_token
        payload["bill_number"] = opd.bill_number
        updated_opd = OPD(opd_id=opd_id, **payload)
        _update_opd_row(updated_opd)
        return redirect(url_for("index_opd", success="OPD record updated successfully."))

    return render_template(
        "edit_opd.html",
        opd=opd,
        sections=OPD_FIELD_SECTIONS,
        form_data=_extract_opd_prefill({}, opd),
        error=None,
        success=None,
    )


@app.route("/opd/view-all", methods=["GET"])
def view_all_opd():
    all_opd = _get_opd_records()
    search_query = request.args.get("search", "").strip()
    
    if search_query:
        lowered = search_query.lower()
        filtered_opd = [
            opd
            for opd in all_opd
            if lowered in str(opd.opd_id).lower()
            or lowered in opd.patient_name.lower()
            or lowered in opd.mobile_number.lower()
            or lowered in opd.opd_token.lower()
            or lowered in opd.bill_number.lower()
            or lowered in (opd.doctor_name or "").lower()
            or lowered in (opd.department or "").lower()
        ]
    else:
        filtered_opd = all_opd
    
    page = request.args.get("page", "1")
    try:
        page = int(page)
        if page < 1:
            page = 1
    except (ValueError, TypeError):
        page = 1
    
    per_page = 20
    total_opd = len(filtered_opd)
    total_pages = (total_opd + per_page - 1) // per_page if total_opd > 0 else 1
    
    if page > total_pages:
        page = total_pages
    
    start_idx = (page - 1) * per_page
    end_idx = start_idx + per_page
    opd_page = filtered_opd[start_idx:end_idx]
    
    return render_template(
        "view_all_opd.html",
        opd_records=opd_page,
        page=page,
        total_pages=total_pages,
        total_opd=total_opd,
        per_page=per_page,
        search_query=search_query,
    )


@app.route("/admissions", methods=["GET"])
def index_admissions():
    all_admissions = _get_admissions()
    search_query = request.args.get("search", "").strip()
    if search_query:
        lowered = search_query.lower()
        search_results = [
            admission
            for admission in all_admissions
            if lowered in str(admission.admission_id).lower()
            or lowered in admission.patient_name.lower()
            or lowered in admission.patient_id.lower()
            or lowered in (admission.admission_id or "").lower()
            or lowered in (admission.room_number or "").lower()
            or lowered in (admission.bed_number or "").lower()
        ]
    else:
        search_results = []

    selected_admission = None
    selected_id = request.args.get("selected_id", "").strip()
    if selected_id:
        try:
            admission_id = int(selected_id)
            selected_admission = _find_admission(admission_id)
            if not selected_admission:
                for admission in all_admissions:
                    if str(admission.admission_id) == selected_id or str(admission.admission_id) == str(admission_id):
                        selected_admission = admission
                        break
        except (ValueError, TypeError):
            selected_admission = None

    form_data = _extract_admission_prefill({}, selected_admission) if selected_admission else _extract_admission_prefill({})

    return render_template(
        "index_admission.html",
        sections=ADMISSION_FIELD_SECTIONS,
        form_data=form_data,
        error=None,
        success=request.args.get("success"),
        search_query=search_query,
        search_results=search_results,
        selected_admission=selected_admission,
    )


@app.post("/admissions")
def create_admission():
    form = request.form
    error = _validate_admission_payload(form)
    if error:
        all_admissions = _get_admissions()
        return (
            render_template(
                "index_admission.html",
                sections=ADMISSION_FIELD_SECTIONS,
                form_data=_extract_admission_prefill(form),
                error=error,
                success=None,
                search_query="",
                search_results=all_admissions[:5],
                selected_admission=None,
            ),
            400,
        )
    payload = _extract_admission_payload(form)
    admission = _create_admission(payload)
    name = admission.patient_name or "Unnamed Patient"
    try:
        adm_id = int(admission.admission_id) if admission.admission_id else 0
        admission_id_display = f"ADM{adm_id:05d}"
    except (ValueError, TypeError):
        admission_id_display = str(admission.admission_id) if admission.admission_id else "N/A"
    success_msg = f"Admission created: {name} | Admission ID: {admission_id_display}"
    return redirect(url_for("index_admissions", success=success_msg))


@app.route("/admissions/<int:admission_id>/edit", methods=["GET", "POST"])
def edit_admission(admission_id: int):
    admission = _find_admission(admission_id)
    if not admission:
        abort(404)

    if request.method == "POST":
        form = request.form
        error = _validate_admission_payload(form)
        if error:
            return (
                render_template(
                    "edit_admission.html",
                    admission=admission,
                    sections=ADMISSION_FIELD_SECTIONS,
                    form_data=_extract_admission_prefill(form, admission),
                    error=error,
                    success=None,
                ),
                400,
            )
        payload = _extract_admission_payload(form)
        updated_admission = Admission(admission_id=admission_id, **payload)
        _update_admission_row(updated_admission)
        return redirect(url_for("index_admissions", success="Admission updated successfully."))

    return render_template(
        "edit_admission.html",
        admission=admission,
        sections=ADMISSION_FIELD_SECTIONS,
        form_data=_extract_admission_prefill({}, admission),
        error=None,
        success=None,
    )


@app.route("/admissions/view-all", methods=["GET"])
def view_all_admissions():
    all_admissions = _get_admissions()
    search_query = request.args.get("search", "").strip()
    
    if search_query:
        lowered = search_query.lower()
        filtered_admissions = [
            admission
            for admission in all_admissions
            if lowered in str(admission.admission_id).lower()
            or lowered in admission.patient_name.lower()
            or lowered in admission.patient_id.lower()
            or lowered in (admission.admission_id or "").lower()
            or lowered in (admission.room_number or "").lower()
            or lowered in (admission.bed_number or "").lower()
            or lowered in (admission.admitting_consultant or "").lower()
            or lowered in (admission.admitting_department or "").lower()
        ]
    else:
        filtered_admissions = all_admissions
    
    page = request.args.get("page", "1")
    try:
        page = int(page)
        if page < 1:
            page = 1
    except (ValueError, TypeError):
        page = 1
    
    per_page = 20
    total_admissions = len(filtered_admissions)
    total_pages = (total_admissions + per_page - 1) // per_page if total_admissions > 0 else 1
    
    if page > total_pages:
        page = total_pages
    
    start_idx = (page - 1) * per_page
    end_idx = start_idx + per_page
    admissions_page = filtered_admissions[start_idx:end_idx]
    
    return render_template(
        "view_all_admissions.html",
        admissions=admissions_page,
        page=page,
        total_pages=total_pages,
        total_admissions=total_admissions,
        per_page=per_page,
        search_query=search_query,
    )


@app.route("/charges", methods=["GET", "POST"])
def index_charges():
    charge_master = _get_charge_master()
    success = request.args.get("success")
    
    if request.method == "POST":
        payload = _extract_charge_payload(request.form)
        charge_master = _save_charge_master(payload)
        success_msg = "Charge Master updated successfully."
        return redirect(url_for("index_charges", success=success_msg))
    
    form_data = _prepare_charge_form_data(charge_master)
    
    return render_template(
        "index_charge.html",
        sections=CHARGE_FIELD_SECTIONS,
        form_data=form_data,
        success=success,
        charge_master=charge_master,
    )


@app.route("/billing", methods=["GET", "POST"])
def index_billing():
    charge_master = _get_charge_master()
    all_patients = _get_patients()
    all_admissions = _get_admissions()
    
    search_query = request.args.get("search", "").strip()
    selected_patient_id = request.args.get("patient_id", type=int)
    selected_admission_id = request.args.get("admission_id", type=int)
    
    selected_patient = None
    selected_admission = None
    
    if selected_patient_id:
        selected_patient = _find_patient(selected_patient_id)
        if selected_patient:
            # Find admission for this patient (get the most recent one)
            matching_admissions = []
            for admission in all_admissions:
                if admission.patient_id and str(admission.patient_id) == str(selected_patient_id):
                    if admission.admission_id and admission.admission_id != "":
                        matching_admissions.append(admission)
            if matching_admissions:
                # Sort by admission_id descending to get the most recent
                matching_admissions.sort(key=lambda x: int(x.admission_id) if x.admission_id else 0, reverse=True)
                selected_admission = matching_admissions[0]
    
    if selected_admission_id:
        selected_admission = _find_admission(selected_admission_id)
        if selected_admission and selected_admission.patient_id:
            try:
                selected_patient = _find_patient(int(selected_admission.patient_id))
            except (ValueError, TypeError):
                pass
    
    if request.method == "POST":
        # Check which button was clicked
        action = request.form.get("action", "generate").strip()  # "save" or "generate"
        
        # Process billing form
        patient_id = request.form.get("patient_id", "").strip()
        admission_id = request.form.get("admission_id", "").strip()
        billing_type = request.form.get("billing_type", "IPD").strip()
        
        # Bills can only be generated for admitted patients
        if not admission_id:
            error_msg = "Billing is only available for admitted patients. Please select an admission."
            return redirect(url_for("index_billing", 
                                  patient_id=patient_id if patient_id else None,
                                  admission_id=admission_id if admission_id else None,
                                  error=error_msg))
        
        # Registration & Administrative charges that can't be duplicated
        registration_charges = {
            "registration_fee", "file_opening_charges", "card_opd_slip_charges",
            "admission_processing_fee", "emergency_registration_fee"
        }
        
        # Room / Bed Charges that can't be duplicated (per day charges)
        # Note: nursing_care_charge is handled separately - it's applied once per day
        room_bed_charges_set = {
            "general_ward_bed", "semi_private_room", "private_room", "deluxe_room", 
            "suite_room", "icu", "iccu", "nicu_picu", "ventilator_bed", "isolation_room"
        }
        
        # Gather existing admission charge data and bill usage info
        billing_state = _collect_admission_billing_state(admission_id)
        existing_bills_for_admission = billing_state["existing_bills"]
        pending_charge_entries = billing_state["pending_charge_entries"]
        used_registration_charges = set(billing_state["used_registration_charges"])
        used_room_bed_charges = set(billing_state["used_room_bed_charges"])
        total_nursing_care_days = int(billing_state["total_nursing_care_days"])
        
        # Check if a final bill already exists for this admission (only for Generate Bill action)
        if action == "generate":
            for bill in existing_bills_for_admission:
                if bill.bill_status == "Final":
                    error_msg = f"A final bill (Bill No: {bill.bill_number}) already exists for this admission. Only one final bill can be generated per admission."
                    return redirect(url_for("index_billing", 
                                          patient_id=patient_id if patient_id else None,
                                          admission_id=admission_id if admission_id else None,
                                          error=error_msg))
        
        # Get charges from form
        charges = []
        charge_master_dict = {}
        if charge_master:
            for field_name in CHARGE_FIELD_ORDER:
                charge_master_dict[field_name] = getattr(charge_master, field_name, "")
        
        # Process each charge field
        duplicate_charges = []
        for field_name in CHARGE_FIELD_ORDER:
            quantity = request.form.get(f"qty_{field_name}", "0").strip()
            try:
                qty = int(quantity) if quantity else 0
                if qty > 0:
                    # Check if this is a registration charge that's already been used
                    if field_name in registration_charges and field_name in used_registration_charges:
                        duplicate_charges.append(field_name.replace("_", " ").title())
                        continue
                    
                    # Check if this is a room/bed charge that's already been used
                    if field_name in room_bed_charges_set and field_name in used_room_bed_charges:
                        duplicate_charges.append(field_name.replace("_", " ").title())
                        continue
                    
                    # Check nursing care charge - should be applied once per day per admission
                    if field_name == "nursing_care_charge" and admission_id:
                        # Get admission to calculate days
                        admission_for_nursing = None
                        if selected_admission:
                            admission_for_nursing = selected_admission
                        else:
                            try:
                                admission_for_nursing = _find_admission(int(admission_id))
                            except (ValueError, TypeError):
                                pass
                        
                        # Calculate available days (admission days - already applied days)
                        admission_days_for_nursing = 0
                        if admission_for_nursing and admission_for_nursing.admission_date_time:
                            try:
                                admission_dt_str = admission_for_nursing.admission_date_time
                                if "T" in admission_dt_str:
                                    admission_dt = datetime.strptime(admission_dt_str, "%Y-%m-%dT%H:%M")
                                else:
                                    admission_dt = datetime.strptime(admission_dt_str.split()[0], "%Y-%m-%d")
                                
                                if admission_for_nursing.discharge_date_time:
                                    discharge_dt_str = admission_for_nursing.discharge_date_time
                                    if "T" in discharge_dt_str:
                                        discharge_dt = datetime.strptime(discharge_dt_str, "%Y-%m-%dT%H:%M")
                                    else:
                                        discharge_dt = datetime.strptime(discharge_dt_str.split()[0], "%Y-%m-%d")
                                    end_dt = discharge_dt
                                else:
                                    end_dt = datetime.now()
                                
                                delta = end_dt - admission_dt
                                admission_days_for_nursing = max(1, delta.days + (1 if delta.seconds > 0 else 0))
                            except (ValueError, AttributeError, TypeError):
                                admission_days_for_nursing = 1
                        
                        # Check if trying to add more days than available
                        available_days = admission_days_for_nursing - total_nursing_care_days
                        if qty > available_days:
                            if available_days <= 0:
                                duplicate_charges.append(f"Nursing Care Charge (already applied for all {admission_days_for_nursing} days)")
                                continue
                            else:
                                # Limit to available days
                                qty = available_days
                    
                    charge_amount = charge_master_dict.get(field_name, "0")
                    try:
                        amount = float(charge_amount) if charge_amount else 0.0
                        charges.append({
                            "charge_name": field_name.replace("_", " ").title(),
                            "charge_code": field_name,
                            "quantity": qty,
                            "unit_price": amount,
                            "total": amount * qty
                        })
                    except (ValueError, TypeError):
                        pass
            except (ValueError, TypeError):
                pass
        
        # If there are duplicate charges, return error
        if duplicate_charges:
            error_msg = f"The following charges have already been applied and cannot be added again: {', '.join(duplicate_charges)}"
            # Re-render form with error - redirect to GET with error message
            return redirect(url_for("index_billing", 
                                  patient_id=patient_id if patient_id else None,
                                  admission_id=admission_id if admission_id else None,
                                  error=error_msg))
        
        # If generating final bill, combine all draft bills and saved charge entries for this admission
        draft_bills: List[Billing] = []
        if action == "generate" and admission_id:
            draft_bills = [bill for bill in existing_bills_for_admission if bill.bill_status == "Draft"]
            combined_charges: Dict[str, Dict[str, Any]] = {}  # key: charge_code
            total_discount = 0.0
            total_tax = 0.0

            def _merge_charge_list(charge_list: List[Dict[str, Any]]):
                for charge in charge_list:
                    charge_code = charge.get("charge_code", "")
                    if not charge_code:
                        continue
                    try:
                        quantity = float(charge.get("quantity", 0))
                    except (ValueError, TypeError):
                        quantity = 0.0
                    try:
                        total_value = float(charge.get("total", 0))
                    except (ValueError, TypeError):
                        total_value = 0.0
                    try:
                        unit_price = float(charge.get("unit_price", 0))
                    except (ValueError, TypeError):
                        unit_price = 0.0
                    if charge_code in combined_charges:
                        existing = combined_charges[charge_code]
                        existing["quantity"] += quantity
                        existing["total"] += total_value
                    else:
                        combined_charges[charge_code] = {
                            "charge_name": charge.get("charge_name", charge_code.replace("_", " ").title()),
                            "charge_code": charge_code,
                            "quantity": quantity,
                            "unit_price": unit_price,
                            "total": total_value,
                        }

            # Include charges from previously saved admission charge entries (pending ones only)
            for entry in pending_charge_entries:
                charges_from_entry = getattr(entry, "parsed_charges", [])
                _merge_charge_list(charges_from_entry)
                try:
                    total_discount += float(entry.discount) if entry.discount else 0.0
                except (ValueError, TypeError):
                    pass
                try:
                    total_tax += float(entry.tax) if entry.tax else 0.0
                except (ValueError, TypeError):
                    pass

            # Include charges from legacy draft bills (if any exist from previous versions)
            for draft_bill in draft_bills:
                if draft_bill.charges_json:
                    try:
                        draft_charges = json.loads(draft_bill.charges_json)
                        _merge_charge_list(draft_charges)
                        total_discount += float(draft_bill.discount) if draft_bill.discount else 0.0
                        total_tax += float(draft_bill.tax) if draft_bill.tax else 0.0
                    except (json.JSONDecodeError, TypeError, ValueError):
                        pass

            # Include charges selected in the current form submission
            _merge_charge_list(charges)

            # Convert combined charges back to list
            charges = []
            for merged in combined_charges.values():
                # Normalize quantity to int when applicable
                qty = merged["quantity"]
                merged["quantity"] = int(qty) if abs(qty - int(qty)) < 1e-6 else qty
                merged["total"] = float(merged["total"])
                charges.append(merged)

            # Add discount and tax from current form
            discount_str = request.form.get("discount", "0").strip()
            discount = float(discount_str) if discount_str else 0.0
            tax_str = request.form.get("tax", "0").strip()
            tax = float(tax_str) if tax_str else 0.0
            total_discount += discount
            total_tax += tax
        else:
            # For save action, just use current form values
            discount_str = request.form.get("discount", "0").strip()
            discount = float(discount_str) if discount_str else 0.0
            tax_str = request.form.get("tax", "0").strip()
            tax = float(tax_str) if tax_str else 0.0
            total_discount = discount
            total_tax = tax

        # Ensure there is at least one charge to process
        if not charges:
            error_msg = "Please select at least one charge before saving or generating a bill."
            return redirect(url_for("index_billing", 
                                  patient_id=patient_id if patient_id else None,
                                  admission_id=admission_id if admission_id else None,
                                  error=error_msg))
        
        # Calculate totals
        subtotal = sum(charge["total"] for charge in charges)
        total_amount = subtotal - total_discount + total_tax
        
        # Get patient name
        patient_name = ""
        if selected_patient:
            patient_name = selected_patient.full_name or ""
        elif patient_id:
            patient = _find_patient(int(patient_id))
            if patient:
                patient_name = patient.full_name or ""
        
        if action == "save":
            entry_payload = {
                "patient_id": patient_id,
                "patient_name": patient_name,
                "admission_id": admission_id,
                "billing_type": billing_type,
                "charges_json": json.dumps(charges),
                "subtotal": str(subtotal),
                "discount": str(total_discount),
                "tax": str(total_tax),
                "total_amount": str(total_amount),
                "status": "Pending",
            }
            new_entry = _create_admission_charge(entry_payload)
            success_msg = f"Charges saved successfully (Entry #CHG{new_entry.charge_entry_id:05d}). Total: ₹{total_amount:.2f}"
            return redirect(url_for("index_billing", patient_id=patient_id, admission_id=admission_id, success=success_msg))
        else:
            payload = {
                "patient_id": patient_id,
                "patient_name": patient_name,
                "admission_id": admission_id,
                "billing_date": datetime.now().strftime("%Y-%m-%d"),
                "billing_type": billing_type,
                "charges_json": json.dumps(charges),
                "subtotal": str(subtotal),
                "discount": str(total_discount),
                "tax": str(total_tax),
                "total_amount": str(total_amount),
                "payment_status": "Pending",
                "payment_mode": "",
                "payment_reference": "",
                "notes": "",
                "bill_status": "Final",
            }
            
            new_bill = _create_bill(payload)
            
            # Mark legacy draft bills as merged
            for bill in draft_bills:
                if bill.bill_id != new_bill.bill_id:
                    bill.bill_status = "Merged"
                    _update_bill_row(bill)
            
            # Mark saved admission charge entries as merged once included in final bill
            for entry in pending_charge_entries:
                entry.status = "Merged"
                _update_admission_charge_row(entry)
            
            return redirect(url_for("view_bill", bill_id=new_bill.bill_id))
    
    # Search results - support both patient search and admission ID search
    search_results = []
    admission_search_results = []
    if search_query:
        lowered = search_query.lower()
        query_upper = search_query.upper()
        
        # Check if search query matches admission ID format (ADM followed by digits)
        admission_id_match = re.match(r'^ADM(\d+)$', query_upper)
        admission_id_number = None
        if admission_id_match:
            try:
                admission_id_number = int(admission_id_match.group(1))
            except (ValueError, TypeError):
                pass
        
        # Search for admissions by ID
        if admission_id_number is not None:
            for admission in all_admissions:
                if admission.admission_id == admission_id_number:
                    admission_search_results.append(admission)
        else:
            # Also search for partial admission ID matches
            for admission in all_admissions:
                try:
                    adm_id = int(admission.admission_id) if admission.admission_id else 0
                    adm_id_str = f"ADM{adm_id:05d}"
                    if lowered in adm_id_str.lower():
                        admission_search_results.append(admission)
                except (ValueError, TypeError):
                    # If admission_id can't be converted, skip this admission
                    pass
        
        # Search for patients
        for patient in all_patients:
            if (lowered in str(patient.patient_id).lower() or
                lowered in patient.full_name.lower() or
                lowered in patient.hospital_id.lower() or
                lowered in patient.mobile_primary.lower()):
                search_results.append(patient)
    
    success = request.args.get("success")
    error = request.args.get("error")
    
    # Calculate number of days for room/bed charges
    admission_days = 0
    # Charges that are auto-calculated per day
    room_bed_charges = {
        "general_ward_bed",
        "nursing_care_charge"
    }
    
    if selected_admission and selected_admission.admission_date_time:
        try:
            # Parse admission date
            admission_dt_str = selected_admission.admission_date_time
            # Handle both "YYYY-MM-DDTHH:MM" and "YYYY-MM-DD HH:MM:SS" formats
            if "T" in admission_dt_str:
                admission_dt = datetime.strptime(admission_dt_str, "%Y-%m-%dT%H:%M")
            else:
                admission_dt = datetime.strptime(admission_dt_str.split()[0], "%Y-%m-%d")
            
            # Use discharge date if available, otherwise use current date
            if selected_admission.discharge_date_time:
                discharge_dt_str = selected_admission.discharge_date_time
                if "T" in discharge_dt_str:
                    discharge_dt = datetime.strptime(discharge_dt_str, "%Y-%m-%dT%H:%M")
                else:
                    discharge_dt = datetime.strptime(discharge_dt_str.split()[0], "%Y-%m-%d")
                end_dt = discharge_dt
            else:
                end_dt = datetime.now()
            
            # Calculate days (at least 1 day even if same day)
            delta = end_dt - admission_dt
            admission_days = max(1, delta.days + (1 if delta.seconds > 0 else 0))
        except (ValueError, AttributeError, TypeError) as e:
            # If date parsing fails, default to 1 day
            admission_days = 1
    
    # Get existing saved charge entries and bills for this admission
    charge_entries = []
    existing_bills = []
    used_registration_charges = set()
    used_room_bed_charges = set()
    total_nursing_care_days = 0
    
    if selected_admission and selected_admission.admission_id:
        billing_state = _collect_admission_billing_state(selected_admission.admission_id)
        charge_entries = billing_state["charge_entries"]
        existing_bills = billing_state["existing_bills"]
        used_registration_charges = set(billing_state["used_registration_charges"])
        used_room_bed_charges = set(billing_state["used_room_bed_charges"])
        total_nursing_care_days = int(billing_state["total_nursing_care_days"])
    
    # Convert charge_master to dict for template access
    charge_master_dict = {}
    if charge_master:
        for field_name in CHARGE_FIELD_ORDER:
            charge_master_dict[field_name] = getattr(charge_master, field_name, "0")
    
    # All room/bed charges for validation (not just auto-calculated ones)
    all_room_bed_charges = {
        "general_ward_bed", "semi_private_room", "private_room", "deluxe_room", 
        "suite_room", "icu", "iccu", "nicu_picu", "ventilator_bed", "isolation_room"
    }
    
    return render_template(
        "index_billing.html",
        charge_master=charge_master_dict,
        charge_field_order=CHARGE_FIELD_ORDER,
        charge_field_sections=CHARGE_FIELD_SECTIONS,
        selected_patient=selected_patient,
        selected_admission=selected_admission,
        search_query=search_query,
        search_results=search_results[:10],
        admission_search_results=admission_search_results[:10],
        charge_entries=charge_entries,
        existing_bills=existing_bills,
        used_registration_charges=list(used_registration_charges),
        used_room_bed_charges=list(used_room_bed_charges),
        total_nursing_care_days=total_nursing_care_days,
        admission_days=admission_days,
        room_bed_charges=list(room_bed_charges),
        all_room_bed_charges=list(all_room_bed_charges),
        success=success,
        error=error,
    )


@app.route("/billing/<int:bill_id>/view", methods=["GET"])
def view_bill(bill_id: int):
    bill = _find_bill(bill_id)
    if not bill:
        abort(404)
    
    # Parse charges JSON
    charges = []
    if bill.charges_json:
        try:
            charges = json.loads(bill.charges_json)
        except (json.JSONDecodeError, TypeError):
            charges = []
    
    # Get patient details if available
    patient = None
    if bill.patient_id:
        try:
            patient = _find_patient(int(bill.patient_id))
        except (ValueError, TypeError):
            pass
    
    # Get admission details if available
    admission = None
    if bill.admission_id:
        try:
            admission = _find_admission(int(bill.admission_id))
        except (ValueError, TypeError):
            pass
    
    return render_template(
        "view_bill.html",
        bill=bill,
        charges=charges,
        patient=patient,
        admission=admission,
    )


@app.route("/billing/<int:bill_id>/edit", methods=["GET", "POST"])
def edit_bill(bill_id: int):
    bill = _find_bill(bill_id)
    if not bill:
        abort(404)
    
    charge_master = _get_charge_master()
    
    # Parse charges JSON
    charges = []
    if bill.charges_json:
        try:
            charges = json.loads(bill.charges_json)
        except (json.JSONDecodeError, TypeError):
            charges = []
    
    # Get patient details if available
    patient = None
    if bill.patient_id:
        try:
            patient = _find_patient(int(bill.patient_id))
        except (ValueError, TypeError):
            pass
    
    # Get admission details if available
    admission = None
    if bill.admission_id:
        try:
            admission = _find_admission(int(bill.admission_id))
        except (ValueError, TypeError):
            pass
    
    if request.method == "POST":
        action = request.form.get("action")
        
        if action == "delete_charge":
            # Delete a specific charge by index
            charge_index = request.form.get("charge_index")
            if charge_index is not None:
                try:
                    index = int(charge_index)
                    if 0 <= index < len(charges):
                        charges.pop(index)
                        # Recalculate bill totals
                        subtotal = sum(float(c.get("total", 0)) for c in charges)
                        discount = float(bill.discount) if bill.discount else 0.0
                        tax = float(bill.tax) if bill.tax else 0.0
                        total_amount = subtotal - discount + tax
                        
                        # Update bill
                        bill.charges_json = json.dumps(charges)
                        bill.subtotal = str(subtotal)
                        bill.total_amount = str(total_amount)
                        _update_bill_row(bill)
                        
                        return redirect(url_for("edit_bill", bill_id=bill_id, success="Charge deleted successfully."))
                except (ValueError, TypeError):
                    pass
        
        elif action == "update_bill":
            # Update charges with new quantities and recalculate
            updated_charges = []
            for i, charge in enumerate(charges):
                charge_code = charge.get("charge_code", "")
                qty_key = f"qty_{charge_code}"
                qty = request.form.get(qty_key)
                if qty:
                    try:
                        qty = float(qty)
                        if qty > 0:
                            charge["quantity"] = qty
                            charge["total"] = float(charge.get("unit_price", 0)) * qty
                            updated_charges.append(charge)
                    except (ValueError, TypeError):
                        pass
            
            # Get discount and tax from form
            discount = request.form.get("discount", "0")
            tax = request.form.get("tax", "0")
            try:
                discount = float(discount)
                tax = float(tax)
            except (ValueError, TypeError):
                discount = float(bill.discount) if bill.discount else 0.0
                tax = float(bill.tax) if bill.tax else 0.0
            
            # Recalculate totals
            subtotal = sum(float(c.get("total", 0)) for c in updated_charges)
            total_amount = subtotal - discount + tax
            
            # Update bill
            bill.charges_json = json.dumps(updated_charges)
            bill.subtotal = str(subtotal)
            bill.discount = str(discount)
            bill.tax = str(tax)
            bill.total_amount = str(total_amount)
            _update_bill_row(bill)
            
            return redirect(url_for("edit_bill", bill_id=bill_id, success="Bill updated successfully."))
        
        elif action == "regenerate_bill":
            # Regenerate final bill (if it's a draft, convert to final)
            if bill.bill_status == "Draft":
                bill.bill_status = "Final"
                _update_bill_row(bill)
                return redirect(url_for("view_bill", bill_id=bill_id))
            else:
                # For final bills, just redirect to view
                return redirect(url_for("view_bill", bill_id=bill_id))
    
    # Build charge field sections for display
    charge_field_sections = []
    if charge_master:
        # Group charges by section
        sections_map = {
            "registration": {"label": "Registration & Administrative Charges", "fields": []},
            "room_bed": {"label": "Room / Bed Charges (IPD) - Per Day", "fields": []},
            "nursing": {"label": "Nursing Charges", "fields": []},
            "doctor_visit": {"label": "Doctor Visit Charges", "fields": []},
            "procedures": {"label": "Procedures & Treatment Charges", "fields": []},
            "ot": {"label": "OT / Procedure Room Charges", "fields": []},
            "pharmacy": {"label": "Pharmacy / Medicine Charges", "fields": []},
            "bedside": {"label": "Bedside Services", "fields": []},
            "misc": {"label": "Miscellaneous Charges", "fields": []},
            "package": {"label": "Package Billing", "fields": []},
        }
        
        # Map charge codes to sections (simplified - you may need to adjust based on your field names)
        for idx, charge in enumerate(charges):
            charge_code = charge.get("charge_code", "")
            charge_name = charge.get("charge_name", "")
            unit_price = float(charge.get("unit_price", 0))
            quantity = float(charge.get("quantity", 0))
            total = float(charge.get("total", 0))
            
            # Determine section based on charge code
            section_key = "misc"
            if "registration" in charge_code or "file_opening" in charge_code or "card_opd" in charge_code or "admission_processing" in charge_code or "emergency_registration" in charge_code:
                section_key = "registration"
            elif "ward" in charge_code or "room" in charge_code or "icu" in charge_code or "bed" in charge_code or "ventilator" in charge_code or "isolation" in charge_code:
                section_key = "room_bed"
            elif "nursing" in charge_code:
                section_key = "nursing"
            elif "consultation" in charge_code or "visit" in charge_code or "surgeon" in charge_code:
                section_key = "doctor_visit"
            elif "dressing" in charge_code or "nebulization" in charge_code or "catheterization" in charge_code or "injection" in charge_code or "iv_fluids" in charge_code or "enema" in charge_code or "transfusion" in charge_code or "plaster" in charge_code or "suturing" in charge_code or "physiotherapy" in charge_code or "dialysis" in charge_code:
                section_key = "procedures"
            elif "ot" in charge_code or "anesthesia" in charge_code or "recovery" in charge_code:
                section_key = "ot"
            elif "tablet" in charge_code or "injection" in charge_code or "consumable" in charge_code:
                section_key = "pharmacy"
            elif "oxygen" in charge_code or "ventilator" in charge_code or "defibrillator" in charge_code or "cpap" in charge_code or "bipap" in charge_code or "suction" in charge_code:
                section_key = "bedside"
            elif "package" in charge_code or "delivery" in charge_code or "cesarean" in charge_code or "cataract" in charge_code or "knee" in charge_code or "covid" in charge_code:
                section_key = "package"
            
            sections_map[section_key]["fields"].append({
                "charge_code": charge_code,
                "charge_name": charge_name,
                "unit_price": unit_price,
                "quantity": quantity,
                "total": total,
                "charge_index": idx,  # Store original index for deletion
            })
        
        # Only include sections that have charges
        for key, section in sections_map.items():
            if section["fields"]:
                charge_field_sections.append({
                    "key": key,
                    "label": section["label"],
                    "fields": section["fields"],
                })
    
    return render_template(
        "edit_bill.html",
        bill=bill,
        charges=charges,
        patient=patient,
        admission=admission,
        charge_master=charge_master,
        charge_field_sections=charge_field_sections,
    )


@app.route("/billing/<int:bill_id>/pdf", methods=["GET"])
def download_bill_pdf(bill_id: int):
    """Generate PDF of the bill"""
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from io import BytesIO
    
    bill = _find_bill(bill_id)
    if not bill:
        abort(404)
    
    # Parse charges JSON
    charges = []
    if bill.charges_json:
        try:
            charges = json.loads(bill.charges_json)
        except (json.JSONDecodeError, TypeError):
            charges = []
    
    # Get patient details
    patient = None
    if bill.patient_id:
        try:
            patient = _find_patient(int(bill.patient_id))
        except (ValueError, TypeError):
            pass
    
    # Create PDF in memory
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#0b3d60'),
        spaceAfter=30,
    )
    elements.append(Paragraph("Hospital Bill", title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Bill Info
    bill_data = [
        ["Bill Number:", bill.bill_number or "N/A"],
        ["Bill Date:", bill.billing_date or "N/A"],
        ["Bill Type:", bill.billing_type or "N/A"],
    ]
    if patient:
        bill_data.append(["Patient Name:", patient.full_name or "N/A"])
        bill_data.append(["MRN:", patient.hospital_id or "N/A"])
        bill_data.append(["Mobile:", patient.mobile_primary or "N/A"])
    
    bill_table = Table(bill_data, colWidths=[2*inch, 4*inch])
    bill_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('BACKGROUND', (1, 0), (1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(bill_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Charges Table
    charge_data = [["Charge", "Quantity", "Unit Price", "Total"]]
    for charge in charges:
        charge_data.append([
            charge.get("charge_name", ""),
            str(charge.get("quantity", 0)),
            f"₹{charge.get('unit_price', 0):.2f}",
            f"₹{charge.get('total', 0):.2f}"
        ])
    
    charge_table = Table(charge_data, colWidths=[3*inch, 1*inch, 1.5*inch, 1.5*inch])
    charge_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0b3d60')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
        ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
    ]))
    elements.append(charge_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Summary
    subtotal = float(bill.subtotal) if bill.subtotal else 0.0
    discount = float(bill.discount) if bill.discount else 0.0
    tax = float(bill.tax) if bill.tax else 0.0
    total = float(bill.total_amount) if bill.total_amount else 0.0
    
    summary_data = [
        ["Subtotal:", f"₹{subtotal:.2f}"],
        ["Discount:", f"₹{discount:.2f}"],
        ["Tax:", f"₹{tax:.2f}"],
        ["Total Amount:", f"₹{total:.2f}"]
    ]
    
    summary_table = Table(summary_data, colWidths=[2*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, -2), 'Helvetica'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -2), 10),
        ('FONTSIZE', (0, -1), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(summary_table)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    return app.response_class(
        buffer.getvalue(),
        mimetype='application/pdf',
        headers={'Content-Disposition': f'attachment; filename=bill_{bill.bill_number}.pdf'}
    )


@app.route("/billing/view-all", methods=["GET"])
def view_all_billing():
    all_bills = _get_bills()
    search_query = request.args.get("search", "").strip()
    
    if search_query:
        lowered = search_query.lower()
        filtered_bills = [
            bill
            for bill in all_bills
            if lowered in bill.bill_number.lower()
            or lowered in bill.patient_name.lower()
            or lowered in bill.patient_id.lower()
            or lowered in str(bill.bill_id).lower()
        ]
    else:
        filtered_bills = all_bills
    
    page = request.args.get("page", "1")
    try:
        page = int(page)
        if page < 1:
            page = 1
    except (ValueError, TypeError):
        page = 1
    
    per_page = 20
    total_bills = len(filtered_bills)
    total_pages = (total_bills + per_page - 1) // per_page if total_bills > 0 else 1
    
    if page > total_pages:
        page = total_pages
    
    start_idx = (page - 1) * per_page
    end_idx = start_idx + per_page
    bills_page = filtered_bills[start_idx:end_idx]
    
    return render_template(
        "view_all_billing.html",
        bills=bills_page,
        page=page,
        total_pages=total_pages,
        total_bills=total_bills,
        per_page=per_page,
        search_query=search_query,
    )


if __name__ == "__main__":
    app.run(debug=True)

